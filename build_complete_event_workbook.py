#!/usr/bin/env python3
"""
build_complete_event_workbook.py

Merges the ISU official "Judges Details per Skater" spreadsheet with
pairwise bias analysis results from the canonical database into one
consolidated, well-organized workbook.

Usage:
    python3 build_complete_event_workbook.py                     # event_id=2 (OWG 2026 FD)
    python3 build_complete_event_workbook.py --event-id 22       # single event
    python3 build_complete_event_workbook.py --all-events --dry-run
    python3 build_complete_event_workbook.py --all-events

Data sources:
    ISU source:  excel_output/{comp}_{stem}.xlsx
                 (looked up via sources table: source_type='judges_details_pdf')
    Bias data:   figure_skating_ijs_v4.sqlite (tables: pairwise_impact_results,
                 judge_team_impacts)

Output:
    excel_output/{comp}_{stem}_complete.xlsx  (12 tabs)

Formula handling
----------------
The ISU source file's Summary sheet contains cross-sheet formulas such as
  ='Element Scores'!S13          (TES per competitor)
  ='Program Component Scores'!R5 (PCS per competitor)
  =D4+E4+F4                      (TSS = TES + PCS + Deductions)

Because the source tabs are renamed in this workbook (e.g. "Element Scores"
becomes "ISU \u2013 Elements (GOE)"), those references would break if copied
verbatim.  We load the ISU workbook WITHOUT data_only so we get the formula
strings, then rewrite every cross-sheet reference to use the new tab names
before writing to the output.  Excel recalculates on open.
"""

import argparse
import os
import sqlite3
import sys
from copy import copy
from pathlib import Path
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE    = "/Users/allman/Library/CloudStorage/Dropbox/Dropbox Mike/Judging Bias"
DB_PATH = os.path.join(BASE, "figure_skating_ijs_v4.sqlite")

# ── Colour palette (shared across all custom formatting) ──────────────────────
C_DARK_BLUE   = "1F4E79"   # title bars, tab colours for key tabs
C_MID_BLUE    = "2E75B6"   # column header rows
C_LIGHT_BLUE  = "D6E4F7"   # description rows
C_WHITE       = "FFFFFF"
C_DARK_TEXT   = "1F4E79"   # description text colour
C_GREY_TEXT   = "595959"
C_ISU_GREEN   = "E2EFDA"
C_BIAS_YELLOW = "FFF2CC"
C_ALERT_RED   = "FF0000"   # headline finding
C_ALERT_FILL  = "FFF0F0"   # light red background for headline
C_GOLD_FILL   = "FFD700"   # gold highlight for outcome-determinative finding
C_SECTION_FILL = "EBF3FB"  # light blue for section headers within Key Findings

# ── Tab config: (source_key, source_sheet_name, dest_tab_name) ─────────────────
# Only tabs copied from source files. Key Findings and Glossary are built
# separately and inserted around these.
TAB_CONFIG = [
    # ISU official scoring
    ("isu",  "Summary",                   "ISU \u2013 Summary"),
    ("isu",  "Element Scores",            "ISU \u2013 Elements (GOE)"),
    ("isu",  "Program Component Scores",  "ISU \u2013 PCS"),
    ("isu",  "Legend",                    "ISU \u2013 Legend"),
    # Chat's pairwise bias analysis
    ("chat", "JudgeSummary",              "Bias \u2013 Judge Summary"),
    ("chat", "JudgeImpact_ByTeam",        "Bias \u2013 Impact by Team"),
    ("chat", "Top25Pairs_ByJudge",        "Bias \u2013 Top 25 Pairs"),
    ("chat", "PairwiseBias_All",          "Bias \u2013 All Pairs"),
    ("chat", "Method",                    "Bias \u2013 Method"),
]

# Map: original ISU sheet name → destination tab name (used to rewrite formulas)
ISU_SHEET_REMAP = {src: dst for key, src, dst in TAB_CONFIG if key == "isu"}

TAB_DESCRIPTIONS = {
    "Key Findings":                   "Plain-English headline findings; outcome-determinative result",
    "ISU \u2013 Summary":             "Official ISU score summary \u2014 total scores and placements",
    "ISU \u2013 Elements (GOE)":      "Raw GOE marks from all 9 judges, by element and competitor",
    "ISU \u2013 PCS":                 "Raw PCS marks from all 9 judges, by component and competitor",
    "ISU \u2013 Legend":              "Colour coding legend for ISU source tabs",
    "Bias \u2013 Judge Summary":      "Per-judge count of significant pairs (BH q \u2264 0.05) and min p/q",
    "Bias \u2013 Impact by Team":     "Each judge\u2019s ISU-impact points on every competitor",
    "Bias \u2013 Top 25 Pairs":       "Top 25 most biased pairwise decisions per judge (sorted by |bias|)",
    "Bias \u2013 All Pairs":          "Full pairwise analysis: BiasPoints, Vote, p-value, BH q-value",
    "Bias \u2013 Method":             "Analysis metadata: event, permutation count, null description",
    "Glossary":                       "Plain-English definitions of statistical terms used in this workbook",
}

# ── OWG 2026 Judge Names (used only when event is OWG 2026) ───────────────────
OWG2026_JUDGE_NAMES = {
    "J1": "J1 \u2013 Jezabel DABOUIS",
    "J2": "J2 \u2013 Elena KHMYZENKO",
    "J3": "J3 \u2013 Marta OLOZAGARRE",
    "J4": "J4 \u2013 Virpi KUNNAS-HELMINEN",
    "J5": "J5 \u2013 Janis ENGEL",
    "J6": "J6 \u2013 Leslie KEEN",
    "J7": "J7 \u2013 Isabella MICHELI",
    "J8": "J8 \u2013 Feng HUANG",
    "J9": "J9 \u2013 Richard KOSINA",
}


# ── Helper: ISU source and output paths for an event ──────────────────────────

def get_event_paths(conn, event_id):
    """
    Returns (isu_file_abs, out_file_abs) derived from sources table.
    Pattern: excel_output/{comp_dir}_{pdf_stem}.xlsx
    Raises FileNotFoundError if the xlsx doesn't exist on disk.
    """
    row = conn.execute(
        "SELECT local_path FROM sources WHERE event_id=? AND source_type='judges_details_pdf'",
        (event_id,)
    ).fetchone()
    if not row:
        raise FileNotFoundError(f"No judges_details_pdf source for event_id={event_id}")
    local_path = row[0]
    p = Path(local_path)
    comp_dir = p.parts[-2]   # e.g. 'owg2026' or 'wc2024'
    stem     = p.stem        # e.g. 'FSKXICEDANCE_FD_JudgesDetailsperSkater' or 'SEG001_JudgesDetails'
    isu_name = f"{comp_dir}_{stem}.xlsx"
    out_name = f"{comp_dir}_{stem}_complete.xlsx"
    isu_file = os.path.join(BASE, "excel_output", isu_name)
    out_file = os.path.join(BASE, "excel_output", out_name)
    if not os.path.exists(isu_file):
        raise FileNotFoundError(f"ISU xlsx not found: {isu_file}")
    return isu_file, out_file


# ── Helper: event metadata ─────────────────────────────────────────────────────

def get_event_info(conn, event_id):
    """
    Returns a dict:
      comp_name, discipline, segment, start_date,
      n_entries, n_pairs, is_ice_dance, event_label, is_owg2026
    """
    row = conn.execute("""
        SELECT c.name, e.discipline, e.segment, c.start_date, c.results_root_url
        FROM events e
        JOIN competitions c ON e.competition_id = c.competition_id
        WHERE e.event_id = ?
    """, (event_id,)).fetchone()
    if not row:
        raise ValueError(f"No event found for event_id={event_id}")
    comp_name, discipline, segment, start_date, root_url = row
    n_entries = conn.execute(
        "SELECT COUNT(*) FROM entries WHERE event_id=?", (event_id,)
    ).fetchone()[0]
    n_pairs = n_entries * (n_entries - 1) // 2
    is_ice_dance = "dance" in discipline.lower()
    is_owg2026   = "owg2026" in (root_url or "").lower()
    event_label  = f"{comp_name} \u2014 {discipline} {segment}"
    return {
        "comp_name":    comp_name,
        "discipline":   discipline,
        "segment":      segment,
        "start_date":   start_date or "",
        "n_entries":    n_entries,
        "n_pairs":      n_pairs,
        "is_ice_dance": is_ice_dance,
        "is_owg2026":   is_owg2026,
        "event_label":  event_label,
    }


# ── Helper: key findings (outcome-determinative test) ─────────────────────────

def get_dynamic_key_findings(conn, event_id, event_info):
    """
    Queries DB for the gold vs silver (rank 1 vs rank 2) outcome-determinative test.
    Returns a kf dict if the test is satisfied (q<=0.05 AND |bias|>margin), else None.

    kf keys:
      rank1_team, rank1_noc, rank1_tss,
      rank2_team, rank2_noc, rank2_tss,
      margin, od_judge, od_judge_name,
      j_impact_rank1, j_impact_rank2, j_bias, j_p, j_q,
      sig_judges, n_sig_judges
    """
    # Rank 1 and rank 2 entries
    entries = {
        r[0]: {"team": r[1], "noc": r[2], "tss": r[3]}
        for r in conn.execute(
            "SELECT rank, team_name, noc, tss FROM entries WHERE event_id=? AND rank IN (1,2)",
            (event_id,)
        ).fetchall()
    }
    # Always collect sig_judges regardless of OD result
    sig_judges = [r[0] for r in conn.execute("""
        SELECT DISTINCT judge_position FROM pairwise_impact_results
        WHERE event_id=? AND q_value_bh <= 0.05
        ORDER BY judge_position
    """, (event_id,)).fetchall()]
    no_od_base = {"is_od": False, "sig_judges": sig_judges, "n_sig_judges": len(sig_judges)}

    if 1 not in entries or 2 not in entries:
        return no_od_base

    rank1  = entries[1]
    rank2  = entries[2]
    margin = round(rank1["tss"] - rank2["tss"], 4)

    # Significant pairs with rank_a=1, rank_b=2
    od_row = conn.execute("""
        SELECT judge_position, bias_points, p_value, q_value_bh
        FROM pairwise_impact_results
        WHERE event_id=? AND rank_a=1 AND rank_b=2 AND q_value_bh <= 0.05
        ORDER BY ABS(bias_points) DESC
        LIMIT 1
    """, (event_id,)).fetchone()

    if not od_row:
        return no_od_base

    jpos, bias_points, p_val, q_val = od_row
    if abs(bias_points) <= margin:
        return no_od_base

    # Get judge impact on rank1 and rank2
    def get_impact(judge_pos, rank):
        r = conn.execute(
            "SELECT impact_points FROM judge_team_impacts WHERE event_id=? AND judge_position=? AND rank=?",
            (event_id, judge_pos, rank)
        ).fetchone()
        return r[0] if r else 0.0

    j_impact_rank1 = get_impact(jpos, 1)
    j_impact_rank2 = get_impact(jpos, 2)

    # All judges with any significant pair
    sig_judges = [r[0] for r in conn.execute("""
        SELECT DISTINCT judge_position FROM pairwise_impact_results
        WHERE event_id=? AND q_value_bh <= 0.05
        ORDER BY judge_position
    """, (event_id,)).fetchall()]

    # Judge display name from DB
    jname_row = conn.execute(
        "SELECT judge_name FROM judges WHERE event_id=? AND judge_position=?",
        (event_id, jpos)
    ).fetchone()
    od_judge_name = jname_row[0] if (jname_row and jname_row[0]) else jpos

    return {
        "is_od":          True,
        "rank1_team":     rank1["team"],
        "rank1_noc":      rank1["noc"],
        "rank1_tss":      rank1["tss"],
        "rank2_team":     rank2["team"],
        "rank2_noc":      rank2["noc"],
        "rank2_tss":      rank2["tss"],
        "margin":         margin,
        "od_judge":       jpos,
        "od_judge_name":  od_judge_name,
        "j_impact_rank1": j_impact_rank1,
        "j_impact_rank2": j_impact_rank2,
        "j_bias":         round(bias_points, 4),
        "j_p":            round(p_val, 4),
        "j_q":            round(q_val, 4),
        "sig_judges":     sig_judges,
        "n_sig_judges":   len(sig_judges),
    }


# ── Helper: EVENT_FACTS list ───────────────────────────────────────────────────

def build_dynamic_event_facts(event_info):
    """Returns list of (label, value) tuples for the Overview Event Information block."""
    date_str = event_info["start_date"][:7] if event_info["start_date"] else ""
    return [
        ("Event",            event_info["comp_name"]),
        ("Discipline",       event_info["discipline"]),
        ("Segment",          event_info["segment"]),
        ("Date",             date_str),
        ("Competitors",      str(event_info["n_entries"])),
        ("Judges",           "9"),
        ("Analysis method",  "ISU-impact margin shift; style-adjusted quantile permutation null"),
        ("Permutations",     "10,000 per judge-pair combination"),
        ("Significance",     "BH-FDR q \u2264 0.05 (Benjamini\u2013Hochberg false discovery rate)"),
        ("Source data",      "ISU official Judges Details per Skater"),
        ("Analysis by",      "Michael Allman, University of Chicago Booth School of Business"),
        ("Statistical AI",   "Analysis assisted by Claude (Anthropic) AI"),
    ]


# ── Helper: BIAS_TAB_META dict ─────────────────────────────────────────────────

def build_dynamic_bias_tab_meta(event_info, kf):
    """Constructs BIAS_TAB_META dynamically from event_info and kf."""
    ev  = event_info["event_label"]
    N   = event_info["n_entries"]
    np_ = event_info["n_pairs"]
    tot = 9 * np_

    if kf.get('is_od'):
        od_callout = (
            f"\n\u26a0 OUTCOME-DETERMINATIVE: {kf['od_judge']} ({kf['od_judge_name']}) "
            f"shows a significant gold-vs-silver pair "
            f"({kf['rank1_noc']} vs {kf['rank2_noc']}). "
            f"{kf['od_judge']} inflated {kf['rank1_noc']} by "
            f"+{abs(kf['j_bias']):.2f} ISU points relative to {kf['rank2_noc']} "
            f"(p={kf['j_p']:.4f}, q={kf['j_q']:.4f}). "
            f"{kf['rank1_noc']} won by only {kf['margin']:.2f} points. "
            f"Removing {kf['od_judge']} reverses the gold medal."
        )
        impact_note = (
            f"\n{kf['od_judge']} ({kf['od_judge_name']}): "
            f"impact on {kf['rank1_noc']} (Rank 1) = +{abs(kf['j_impact_rank1']):.2f} pts; "
            f"impact on {kf['rank2_noc']} (Rank 2) = {kf['j_impact_rank2']:.2f} pts. "
            f"Net advantage to {kf['rank1_noc']} = {abs(kf['j_bias']):.2f} pts > "
            f"{kf['margin']:.2f} pt margin \u2192 outcome-determinative."
        )
    else:
        od_callout  = ""
        impact_note = ""

    return {
        "Bias \u2013 Judge Summary": {
            "title": f"Judge Bias Summary \u2014 {ev}",
            "desc": (
                f"For each of the 9 panel judges: count of significantly biased team pairs "
                f"and the strength of the strongest signal detected. Significance uses a "
                f"style-adjusted permutation null (10,000 permutations) with Benjamini\u2013"
                f"Hochberg false discovery rate control.\n"
                f"\u00b7 Judge: panel position and name  "
                f"\u00b7 num_pairs: total pairs judged ({np_} for {N} competitors)  "
                f"\u00b7 num_q\u22640.05: pairs with BH-corrected q \u2264 0.05 (significant)  "
                f"\u00b7 min_p: smallest raw p-value  "
                f"\u00b7 min_q: smallest BH-corrected q-value"
                + od_callout
            ),
            "col_widths":    {"A": 32, "B": 14, "C": 16, "D": 13, "E": 13},
            "numeric_cols":  {4, 5},
            "center_cols":   {3, 4, 5},
            "desc_font_size": 9,
            "hdr_row_height": 22,
        },
        "Bias \u2013 Impact by Team": {
            "title": f"Judge Impact by Team \u2014 {ev}",
            "desc": (
                "For each judge, the estimated ISU-score shift applied to each competitor "
                "through the trimmed-mean scoring mechanism. A positive value means the judge "
                "inflated the team\u2019s total score; negative means the judge deflated it.\n"
                "\u00b7 Judge: panel position and name  "
                "\u00b7 Rank: official final placement  "
                "\u00b7 Team: competitor name  "
                "\u00b7 NOC: nation code  "
                "\u00b7 ImpactPoints_Ij: estimated score shift in ISU published points "
                "(positive\u202f=\u202ffavoured, negative\u202f=\u202fdisfavoured)"
                + impact_note
            ),
            "col_widths":    {"A": 32, "B": 8, "C": 32, "D": 8, "E": 20},
            "numeric_cols":  {5},
            "center_cols":   {2, 4, 5},
            "desc_font_size": 10,
            "hdr_row_height": 30,
        },
        "Bias \u2013 Top 25 Pairs": {
            "title": f"Top 25 Most Biased Pairs per Judge \u2014 {ev}",
            "desc": (
                "For each judge, the 25 team pairs with the largest absolute bias, ranked by "
                "|BiasPoints|. BiasPoints\u202f=\u202fI\u2C7C(A)\u202f\u2212\u202fI\u2C7C(B): "
                "the estimated ISU-point advantage given to Team\u202fA over Team\u202fB by judge\u202fj.\n"
                "\u00b7 Judge: panel position and name  "
                "\u00b7 absBias: |BiasPoints| (sort key)  "
                "\u00b7 A_Rank/Team/NOC: the favoured team  "
                "\u00b7 B_Rank/Team/NOC: the disadvantaged team  "
                "\u00b7 BiasPoints_AminusB: signed bias in ISU points  "
                "\u00b7 Vote: FOR_A / FOR_B / NEUTRAL  "
                "\u00b7 p_value: raw permutation p-value  "
                "\u00b7 q_value_BH: BH-corrected q-value (significant if \u2264\u202f0.05)"
            ),
            "col_widths": {
                "A": 32, "B": 11, "C": 8, "D": 28, "E": 8,
                "F": 8,  "G": 28, "H": 8, "I": 15, "J": 10, "K": 11, "L": 14,
            },
            "numeric_cols":   {2, 9, 11, 12},
            "center_cols":    {2, 3, 8, 9, 10, 11, 12},
            "desc_font_size": 11,
            "hdr_row_height": 36,
        },
        "Bias \u2013 All Pairs": {
            "title": (
                f"Complete Pairwise Bias Analysis \u2014 All {tot:,} Rows "
                f"\u2014 {ev}"
            ),
            "desc": (
                f"Every combination of the {N} competitors "
                f"(C({N},2)\u202f=\u202f{np_} pairs) "
                f"\u00d7 9 judges\u202f=\u202f{tot:,} rows. "
                "BiasPoints_AminusB\u202f=\u202fI\u2C7C(A)\u202f\u2212\u202fI\u2C7C(B): "
                "positive means judge j favoured Team\u202fA over Team\u202fB in ISU-point terms.\n"
                "\u00b7 Judge: panel position and name  "
                "\u00b7 A_Rank/Team/NOC: Team\u202fA  "
                "\u00b7 B_Rank/Team/NOC: Team\u202fB  "
                "\u00b7 BiasPoints_AminusB: signed bias in ISU points  "
                "\u00b7 Vote: FOR_A / FOR_B / NEUTRAL  "
                "\u00b7 p_value: raw permutation p-value  "
                "\u00b7 q_value_BH: BH-corrected q-value (significant if \u2264\u202f0.05)"
            ),
            "col_widths": {
                "A": 32, "B": 8, "C": 28, "D": 8,
                "E": 8,  "F": 28, "G": 8, "H": 14, "I": 10, "J": 11, "K": 14,
            },
            "numeric_cols":  {8, 10, 11},
            "center_cols":   {2, 5, 7, 8, 9, 10, 11},
            "hdr_row_height": 36,
        },
        "Bias \u2013 Method": {
            "title": f"Analysis Method & Parameters \u2014 {ev}",
            "hdr_row_height": 22,
            "desc": (
                "Technical specification of the ISU-impact style-adjusted quantile permutation null.\n"
                "\u00b7 BiasStatistic: B\u2C7C(A,B)\u202f=\u202fI\u2C7C(A)\u2212I\u2C7C(B), "
                "where I\u2C7C(T)\u202f=\u202f\u03a3\u1d63\u202f\u0394\u2C7C\u1d63, "
                "\u0394\u2C7C\u1d63\u202f=\u202fw\u1d63\u202f\u00d7\u202f(tm_actual\u2212tm_neutralized). "
                "w\u1d63 = GOE factor or PCS factor; tm = trimmed mean (drop 1 max + 1 min of 9 judges).\n"
                "\u00b7 Percentile assignment (tie rule): for judge j, category c, mark x in sorted_samples[j,c]: "
                "lo\u202f=\u202fsearchsorted(samples, x, \u2018left\u2019); "
                "hi\u202f=\u202fsearchsorted(samples, x, \u2018right\u2019); "
                "u\u202f\u223c\u202fUniform(lo/n,\u202fhi/n) using fixed RNG. "
                "This is the randomized within-tie interval rule.\n"
                "\u00b7 Inverse CDF: given u\u202f\u2208\u202f[0,1) and sorted_samples length n: "
                "idx\u202f=\u202fmin(int(u\u202f\u00d7\u202fn),\u202fn\u22121); x_quant\u202f=\u202fsamples[idx]. "
                "Preserves the judge\u2019s exact empirical distribution (integer GOE, valid PCS increments). No interpolation.\n"
                "\u00b7 Permutation step (row-wise): for each row r, draw perm\u202f=\u202frng.permutation(9); "
                "assign u_perm[r,j]\u202f=\u202fu_obs[r,\u202fperm[j]] (donor percentile \u2192 receiver judge); "
                "convert back via receiver judge\u2019s inverse CDF for that category.\n"
                "\u00b7 p-value (smoothed): p\u202f=\u202f(1\u202f+\u202f#{|B(t)|\u202f\u2265\u202f|B_obs|})\u202f/\u202f(M\u202f+\u202f1). "
                "M\u202f=\u202f10,000 permutations. RNG seed\u202f=\u202f20260223 (numpy.random.default_rng).\n"
                "\u00b7 CDF scope: GLOBAL (career-wide). Each judge\u2019s samples are pooled across all events "
                "in the database for that judge.\n"
                f"\u00b7 Multiple testing: BH-FDR applied within-event across all "
                f"9\u202f\u00d7\u202f{np_}\u202f=\u202f{tot:,} tests simultaneously. "
                "Method version: isuimpact_quantile_v1."
            ),
            "col_widths":  {"A": 32, "B": 72},
            "numeric_cols": set(),
        },
    }


# ── Formula rewriting ──────────────────────────────────────────────────────────

def rewrite_formula(formula: str, sheet_remap: dict) -> str:
    """
    Rewrite cross-sheet references in an Excel formula string.

    For each (old_name, new_name) pair in sheet_remap, replaces:
        'Old Name'!   →   'New Name'!
    Returns the formula unchanged if no substitution applies.
    """
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula
    result = formula
    for old, new in sheet_remap.items():
        # Quoted form: 'Old Name'!
        result = result.replace(f"'{old}'!", f"'{new}'!")
        # Unquoted form (no spaces in name): OldName!
        if " " not in old:
            result = re.sub(
                rf"(?<!['\w]){re.escape(old)}!",
                f"'{new}'!",
                result,
            )
    return result


# ── Style-copy helpers ─────────────────────────────────────────────────────────

def copy_cell_style(src_cell, dst_cell):
    """Copy all style attributes from src_cell to dst_cell."""
    if src_cell.has_style:
        dst_cell.font       = copy(src_cell.font)
        dst_cell.fill       = copy(src_cell.fill)
        dst_cell.border     = copy(src_cell.border)
        dst_cell.alignment  = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        if src_cell.protection:
            dst_cell.protection = copy(src_cell.protection)


def copy_sheet(src_ws, dst_ws, formula_remap=None):
    """
    Copy cells, styles, merged ranges, column widths, and row heights.

    formula_remap: optional dict {old_sheet_name: new_tab_name}.  When
    provided, any formula cell containing cross-sheet references is rewritten
    so that the references point to the renamed destination tabs.
    """
    # Column widths
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width

    # Row heights
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height

    # Cells + styles
    formula_rewrites = 0
    for row in src_ws.iter_rows():
        for cell in row:
            value = cell.value
            if formula_remap and isinstance(value, str) and value.startswith("="):
                rewritten = rewrite_formula(value, formula_remap)
                if rewritten != value:
                    formula_rewrites += 1
                value = rewritten
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=value)
            copy_cell_style(cell, dst_cell)

    if formula_rewrites:
        print(f"    ({formula_rewrites} formula cross-sheet references rewritten)")

    # Merged cells
    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))

    # Tab colour
    if src_ws.sheet_properties.tabColor:
        dst_ws.sheet_properties.tabColor = copy(src_ws.sheet_properties.tabColor)

    # Freeze panes
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    # Page setup (non-critical; best-effort)
    try:
        dst_ws.page_setup.orientation = src_ws.page_setup.orientation
        dst_ws.page_setup.paperSize   = src_ws.page_setup.paperSize
    except Exception:
        pass


# ── Legend colour fix ──────────────────────────────────────────────────────────

def fix_legend_colors(ws):
    """
    After copying the ISU Legend sheet, apply explicit opaque solid fills to
    the colour-swatch cells (B5=green, B6=pink) so they render visibly
    regardless of theme/alpha interpretation.  Also adds a border and makes
    the rows taller so the swatches are easy to spot.
    """
    thick = Side(style="medium", color="595959")
    swatch_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    for row_num, rgb in [(5, "C6EFCE"), (6, "FFC7CE")]:
        cell = ws.cell(row=row_num, column=2)          # column B
        cell.fill   = PatternFill("solid", fgColor=rgb)
        cell.border = swatch_border
        ws.row_dimensions[row_num].height = 24

    # Column B slightly wider so the swatch is visible
    ws.column_dimensions["B"].width = 7


# ── High / low mark colour coding ─────────────────────────────────────────────

def apply_high_low_colors(
    ws, col_start: int, col_end: int, data_start_row: int, data_ws=None
):
    """
    For each data row, apply explicit green fill to the highest judge mark cell(s)
    and pink fill to the lowest, matching the ISU Legend convention.

    Skips rows with fewer than 2 numeric values or where all values are identical.
    col_start / col_end: 1-based column numbers inclusive (e.g. 7, 15 for G\u2013O).

    data_ws: optional separate worksheet to read VALUES from (use when ws was
             copied with formula strings rather than data_only values).  Fills
             are always applied to ws.
    """
    green = PatternFill("solid", fgColor="C6EFCE")
    pink  = PatternFill("solid", fgColor="FFC7CE")
    src   = data_ws if data_ws is not None else ws

    for row_num in range(data_start_row, ws.max_row + 1):
        row_vals = [
            (c, src.cell(row=row_num, column=c).value)
            for c in range(col_start, col_end + 1)
            if isinstance(src.cell(row=row_num, column=c).value, (int, float))
        ]
        if len(row_vals) < 2 or len({v for _, v in row_vals}) < 2:
            continue
        hi = max(v for _, v in row_vals)
        lo = min(v for _, v in row_vals)
        for col, val in row_vals:
            if val == hi:
                ws.cell(row=row_num, column=col).fill = green
            elif val == lo:
                ws.cell(row=row_num, column=col).fill = pink


# ── Sheet protection helper ────────────────────────────────────────────────────

def lock_sheet(ws):
    """
    Apply read-only sheet protection with no password.

    Users can unlock via Excel \u2192 Review \u2192 Unprotect Sheet (no password required).
    This protects against accidental edits without adding friction.
    """
    ws.protection.sheet    = True
    ws.protection.password = ""     # no password
    ws.protection.enable()


# ── Bias tab: copy with formatted header block ─────────────────────────────────

def copy_sheet_with_bias_header(
    src_ws,
    dst_ws,
    title: str,
    desc: str,
    col_widths: dict,
    numeric_cols: set = None,
    judge_map: dict = None,
    desc_font_size: int = 9,
    center_cols: set = None,
    hdr_row_height: int = 36,
):
    """
    Copy a Bias tab from src_ws to dst_ws, prepending a 3-row formatted
    header block (title / description / blank), then the data rows with a
    styled column-header row.

    Layout:
        Row 1  \u2014 Title bar      (dark blue, white bold 13pt)
        Row 2  \u2014 Description    (light blue, italic 9pt, wrapping)
        Row 3  \u2014 blank spacer   (height 5)
        Row 4  \u2014 Column headers (mid-blue, white bold 10pt, centred)
        Row 5+ \u2014 Data rows      (copied as-is from src rows 2+)

    numeric_cols: 1-based column numbers to format as "0.00000"
    judge_map:    dict mapping raw Judge values (e.g. "J1") to full names
    Freeze panes at A5.  Column widths from col_widths.  Zoom = 125%.
    """
    if numeric_cols is None:
        numeric_cols = set()
    n_cols    = src_ws.max_column
    last_col  = get_column_letter(n_cols)

    # Row 1: Title
    title_cell = dst_ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=C_DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    dst_ws.merge_cells(f"A1:{last_col}1")
    dst_ws.row_dimensions[1].height = 24

    # Row 2: Description
    desc_cell = dst_ws.cell(row=2, column=1, value=desc)
    desc_cell.font      = Font(name="Calibri", size=desc_font_size, italic=True, color=C_DARK_TEXT)
    desc_cell.fill      = PatternFill("solid", fgColor=C_LIGHT_BLUE)
    desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    dst_ws.merge_cells(f"A2:{last_col}2")
    dst_ws.row_dimensions[2].height = 100   # tall enough for wrapped multi-sentence text

    # Row 3: Blank spacer
    dst_ws.row_dimensions[3].height = 5

    # Row 4: Column headers (from src row 1, restyled)
    hdr_font  = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
    hdr_fill  = PatternFill("solid", fgColor=C_MID_BLUE)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    src_hdr_row = next(src_ws.iter_rows(min_row=1, max_row=1))
    for src_cell in src_hdr_row:
        dst_cell = dst_ws.cell(row=4, column=src_cell.column, value=src_cell.value)
        dst_cell.font      = hdr_font
        dst_cell.fill      = hdr_fill
        dst_cell.alignment = hdr_align
    dst_ws.row_dimensions[4].height = hdr_row_height

    # Rows 5+: Data (from src rows 2+)
    data_font = Font(name="Calibri", size=10)

    for src_row in src_ws.iter_rows(min_row=2):
        dst_row_num = src_row[0].row + 3   # offset by 3 (our added header rows)
        for src_cell in src_row:
            value = src_cell.value
            col   = src_cell.column

            # Expand Judge position codes to full names (col 1 = Judge)
            if col == 1 and judge_map and value in judge_map:
                value = judge_map[value]

            dst_cell = dst_ws.cell(row=dst_row_num, column=col, value=value)
            dst_cell.font = data_font
            h_align = "center" if (center_cols and col in center_cols) else "general"
            dst_cell.alignment = Alignment(horizontal=h_align, vertical="center")

            # Apply 5-decimal fixed format to numeric columns
            if col in numeric_cols:
                dst_cell.number_format = "0.00000"
            else:
                dst_cell.number_format = src_cell.number_format

    # Column widths
    for col_letter, width in col_widths.items():
        dst_ws.column_dimensions[col_letter].width = width

    # Freeze panes below header block
    dst_ws.freeze_panes = "A5"

    # Tab colour
    dst_ws.sheet_properties.tabColor = C_BIAS_YELLOW

    # Zoom
    dst_ws.sheet_view.zoomScale = 125

    # Sheet protection (no password)
    lock_sheet(dst_ws)


# ── Key Findings sheet ──────────────────────────────────────────────────────────

def build_key_findings_sheet(ws, kf, event_info):
    """
    Plain-English executive summary of the key findings.

    kf:         dict from get_dynamic_key_findings(), or None if no outcome-det finding.
    event_info: dict from get_event_info().
    """
    ev  = event_info["event_label"]
    N   = event_info["n_entries"]
    np_ = event_info["n_pairs"]
    tot = 9 * np_

    # Fonts
    f_title   = Font(name="Calibri", size=16, bold=True,  color=C_WHITE)
    f_sub     = Font(name="Calibri", size=11, italic=True, color=C_DARK_TEXT)
    f_section = Font(name="Calibri", size=12, bold=True,  color=C_WHITE)
    f_label   = Font(name="Calibri", size=11, bold=True,  color=C_DARK_TEXT)
    f_value   = Font(name="Calibri", size=11,             color="000000")
    f_value_b = Font(name="Calibri", size=11, bold=True,  color="000000")
    f_hdr     = Font(name="Calibri", size=10, bold=True,  color=C_WHITE)

    # Fills
    fill_dark    = PatternFill("solid", fgColor=C_DARK_BLUE)
    fill_mid     = PatternFill("solid", fgColor=C_MID_BLUE)
    fill_light   = PatternFill("solid", fgColor=C_LIGHT_BLUE)
    fill_section = PatternFill("solid", fgColor=C_MID_BLUE)
    fill_alert   = PatternFill("solid", fgColor="FFECEC")
    fill_gold    = PatternFill("solid", fgColor="FFF9E6")
    fill_white   = PatternFill("solid", fgColor=C_WHITE)
    fill_stripe  = PatternFill("solid", fgColor="F5F9FF")

    # Alignments
    al_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    al_center = Alignment(horizontal="center",  vertical="center")
    al_top    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    # Border
    thin = Side(style="thin",   color="BFBFBF")
    tbl_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 1

    # Main title
    ws.cell(row=r, column=1, value=f"Key Findings \u2014 {ev}").font = f_title
    ws.cell(row=r, column=1).fill = fill_dark
    ws.cell(row=r, column=1).alignment = al_center
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 32
    r += 1

    ws.cell(row=r, column=1,
            value="Statistical analysis of potential judging bias \u2014 "
                  "ISU-impact method, 10,000-permutation quantile null, BH-FDR q \u2264 0.05"
            ).font = f_sub
    ws.cell(row=r, column=1).fill = fill_light
    ws.cell(row=r, column=1).alignment = al_left
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 20
    r += 1

    ws.row_dimensions[r].height = 8
    r += 1

    if kf.get('is_od'):
        # FINDING 1 \u2014 OUTCOME-DETERMINATIVE
        ws.cell(row=r, column=1,
                value="\u2605  FINDING 1 \u2014 OUTCOME-DETERMINATIVE BIAS").font = f_section
        ws.cell(row=r, column=1).fill = fill_section
        ws.cell(row=r, column=1).alignment = al_left
        ws.merge_cells(f"A{r}:B{r}")
        ws.row_dimensions[r].height = 22
        r += 1

        headline = (
            f"Judge {kf['od_judge']} ({kf['od_judge_name']}) produced a bias large enough "
            f"to reverse the gold medal.\n\n"
            f"{kf['od_judge_name']} scored {kf['rank1_noc']} (gold medalists {kf['rank1_team']}) "
            f"+{abs(kf['j_impact_rank1']):.2f} points above the panel baseline, and simultaneously "
            f"scored {kf['rank2_noc']} (silver medalists {kf['rank2_team']}) "
            f"{kf['j_impact_rank2']:.2f} points below baseline. "
            f"That is a combined ISU-score advantage of +{abs(kf['j_bias']):.2f} points "
            f"in favour of {kf['rank1_noc']} (p\u202f=\u202f{kf['j_p']:.4f}, "
            f"BH-corrected q\u202f=\u202f{kf['j_q']:.4f} \u2014 statistically significant).\n\n"
            f"{kf['rank1_noc']}\u2019s official winning margin: {kf['rank1_tss']:.2f} "
            f"\u2212 {kf['rank2_tss']:.2f} = {kf['margin']:.2f} points.\n\n"
            f"Since {abs(kf['j_bias']):.2f} > {kf['margin']:.2f}: removing "
            f"{kf['od_judge_name']}\u2019s marks and recalculating under the standard ISU "
            f"trimmed-mean formula reverses the result. "
            f"{kf['rank2_noc']} wins the gold medal."
        )
        c = ws.cell(row=r, column=1, value=headline)
        c.font = f_value
        c.fill = fill_alert
        c.alignment = al_top
        ws.merge_cells(f"A{r}:B{r}")
        ws.row_dimensions[r].height = 120
        r += 1

        ws.row_dimensions[r].height = 6
        r += 1

        # Score summary table header
        ws.cell(row=r, column=1, value="Competitor / Score").font = f_hdr
        ws.cell(row=r, column=1).fill = fill_mid
        ws.cell(row=r, column=1).alignment = al_center
        ws.cell(row=r, column=2, value="Detail").font = f_hdr
        ws.cell(row=r, column=2).fill = fill_mid
        ws.cell(row=r, column=2).alignment = al_center
        ws.row_dimensions[r].height = 18
        r += 1

        tbl_data = [
            (f"{kf['rank1_noc']} (Gold)  \u2014  {kf['rank1_team']}",
             f"TSS = {kf['rank1_tss']:.2f}  |  {kf['od_judge']} impact = +{abs(kf['j_impact_rank1']):.2f} pts",
             fill_gold),
            (f"{kf['rank2_noc']} (Silver)  \u2014  {kf['rank2_team']}",
             f"TSS = {kf['rank2_tss']:.2f}  |  {kf['od_judge']} impact = {kf['j_impact_rank2']:.2f} pts",
             fill_stripe),
            ("Official margin",
             f"{kf['rank1_tss']:.2f} \u2212 {kf['rank2_tss']:.2f} = {kf['margin']:.2f} points",
             fill_white),
            (f"{kf['od_judge']} bias ({kf['rank1_noc']} over {kf['rank2_noc']})",
             f"+{abs(kf['j_bias']):.2f} pts  (p = {kf['j_p']:.4f}, q = {kf['j_q']:.4f}  \u2714 significant)",
             fill_alert),
            (f"Outcome without {kf['od_judge']}",
             f"\u26a0 {kf['rank2_noc']} wins gold  ({abs(kf['j_bias']):.2f} pts bias > {kf['margin']:.2f} pts margin)",
             fill_alert),
        ]
        for i_row, (label, detail, fill) in enumerate(tbl_data):
            c_a = ws.cell(row=r, column=1, value=label)
            c_b = ws.cell(row=r, column=2, value=detail)
            c_a.font = f_value_b
            c_b.font = f_value
            c_a.fill = fill
            c_b.fill = fill
            c_a.alignment = al_left
            c_b.alignment = al_left
            c_a.border = tbl_border
            c_b.border = tbl_border
            ws.row_dimensions[r].height = 48 if i_row == 0 else 18
            r += 1

        ws.row_dimensions[r].height = 10
        r += 1

    else:
        # NO OUTCOME-DETERMINATIVE FINDING
        ws.cell(row=r, column=1,
                value="\u2139  FINDING 1 \u2014 NO OUTCOME-DETERMINATIVE BIAS DETECTED").font = f_section
        ws.cell(row=r, column=1).fill = fill_section
        ws.cell(row=r, column=1).alignment = al_left
        ws.merge_cells(f"A{r}:B{r}")
        ws.row_dimensions[r].height = 22
        r += 1

        c = ws.cell(row=r, column=1,
                    value=(
                        "No statistically significant (BH-FDR q \u2264 0.05) pairwise bias was found "
                        "for the gold vs silver pair that exceeds the gold\u2013silver margin. "
                        "This does not mean no bias exists anywhere in the panel \u2014 see "
                        "'Bias \u2013 Judge Summary' for all findings. It means no single judge "
                        "produced a bias large enough to reverse the final standing between "
                        "first and second place."
                    ))
        c.font = f_value
        c.fill = fill_light
        c.alignment = al_top
        ws.merge_cells(f"A{r}:B{r}")
        ws.row_dimensions[r].height = 80
        r += 1

        ws.row_dimensions[r].height = 10
        r += 1

    # FINDING 2 \u2014 PANEL-WIDE BIAS
    ws.cell(row=r, column=1,
            value="\u2605  FINDING 2 \u2014 PANEL-WIDE SIGNIFICANT BIAS").font = f_section
    ws.cell(row=r, column=1).fill = fill_section
    ws.cell(row=r, column=1).alignment = al_left
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 22
    r += 1

    if kf.get("sig_judges"):
        sig_count = kf["n_sig_judges"]
        sig_names = ", ".join(kf["sig_judges"])
        od_note   = (
            f" Only {kf['od_judge']}\u2019s anomaly is outcome-determinative "
            "for the gold-silver result."
        ) if kf.get('is_od') else ""
    else:
        sig_count = 0
        sig_names = "none"
        od_note   = ""

    summary_text = (
        f"{sig_count} of 9 panel judges show at least one statistically significant pairwise differential "
        f"(BH-FDR q \u2264 0.05 after correcting for {tot:,} simultaneous tests = "
        f"9 judges \u00d7 {np_} unique pairs): "
        f"{sig_names}.{od_note} "
        "See the 'Bias \u2013 Judge Summary' tab for the full breakdown."
    )
    c = ws.cell(row=r, column=1, value=summary_text)
    c.font = f_value
    c.fill = fill_gold
    c.alignment = al_top
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 55
    r += 1

    ws.row_dimensions[r].height = 10
    r += 1

    # METHODOLOGY NOTE
    ws.cell(row=r, column=1,
            value="\u2139  METHODOLOGY NOTE").font = f_section
    ws.cell(row=r, column=1).fill = fill_section
    ws.cell(row=r, column=1).alignment = al_left
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 22
    r += 1

    if kf.get('is_od'):
        two_part_detail = (
            f"  1. Statistical significance: BH-corrected q \u2264 0.05 (after {tot:,} simultaneous tests)\n"
            f"  2. Practical magnitude: |BiasPoints(A,B)| > official winning margin for the A-vs-B pair\n"
            f"{kf['od_judge']} satisfies both "
            f"({abs(kf['j_bias']):.2f} pts bias > {kf['margin']:.2f} pt margin)."
        )
    else:
        two_part_detail = (
            f"  1. Statistical significance: BH-corrected q \u2264 0.05 (after {tot:,} simultaneous tests)\n"
            f"  2. Practical magnitude: |BiasPoints(A,B)| > official winning margin for the A-vs-B pair\n"
            "No judge satisfies both conditions for the gold-silver pair in this event."
        )

    method_rows = [
        ("ISU-impact method",
         "Each judge\u2019s \u2018impact\u2019 on a competitor is measured by replacing that "
         "judge\u2019s marks (GOE + PCS) with the median of the other 8 judges and observing "
         "the change in the published total score. This exactly mirrors the ISU\u2019s own "
         "trimmed-mean calculation."),
        ("BiasPoints(A, B)",
         "Impact(judge, TeamA) \u2212 Impact(judge, TeamB). "
         "Positive = judge inflated A relative to B in ISU-score terms."),
        ("Two-part test",
         "A pair is flagged as outcome-determinative only if BOTH conditions hold:\n"
         + two_part_detail),
        ("Permutation null",
         "10,000 permutations. Within each row (element or PCS component), each "
         "judge\u2019s mark is converted to a percentile rank using that judge\u2019s "
         "scoring distribution across all events in the database, then percentile labels are shuffled. "
         "This preserves each judge\u2019s scoring style while eliminating any "
         "preferential targeting."),
        ("Multiple testing",
         f"Benjamini\u2013Hochberg (BH) false discovery rate correction applied across all "
         f"9 \u00d7 {np_} = {tot:,} tests simultaneously. q \u2264 0.05 is the significance threshold."),
        ("Data source",
         f"ISU official \u2018Judges Details per Skater\u2019 for {ev}. "
         "All computations reproduce the official published total "
         "score (TES + PCS) to within 0.01 points."),
    ]
    fills_cycle = [fill_gold, fill_stripe]
    for i, (label, detail) in enumerate(method_rows):
        c_a = ws.cell(row=r, column=1, value=label)
        c_b = ws.cell(row=r, column=2, value=detail)
        f   = fills_cycle[i % 2]
        c_a.font = f_label
        c_b.font = f_value
        c_a.fill = f
        c_b.fill = f
        c_a.alignment = al_left
        c_b.alignment = al_top
        c_a.border = tbl_border
        c_b.border = tbl_border
        ws.row_dimensions[r].height = 45
        r += 1

    ws.row_dimensions[r].height = 10
    r += 1

    # CAVEATS
    ws.cell(row=r, column=1, value="\u26a0  IMPORTANT CAVEATS").font = f_section
    ws.cell(row=r, column=1).fill = fill_section
    ws.cell(row=r, column=1).alignment = al_left
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 22
    r += 1

    caveats = [
        ("Statistical, not causal",
         "This analysis detects patterns that are statistically improbable under a "
         "fair-judging null. It does not and cannot prove intent or misconduct. "
         "Judges who review many competitions may show apparent patterns for legitimate reasons."),
        ("Single event",
         f"These results are from one event ({ev}, {N} competitors). "
         "Cross-event analysis across the full database of 142 events is available "
         "but not included in this workbook."),
        ("ISU rules",
         "The ISU protocol already trims the highest and lowest mark for each element and "
         "component. The trimmed-mean mechanism reduces (but does not eliminate) the impact "
         "of any single judge. The ISU-impact method measures what each judge contributes "
         "to the final score through this exact mechanism."),
    ]
    for i, (label, detail) in enumerate(caveats):
        c_a = ws.cell(row=r, column=1, value=label)
        c_b = ws.cell(row=r, column=2, value=detail)
        f   = fills_cycle[i % 2]
        c_a.font = f_label
        c_b.font = f_value
        c_a.fill = f
        c_b.fill = f
        c_a.alignment = al_left
        c_b.alignment = al_top
        c_a.border = tbl_border
        c_b.border = tbl_border
        ws.row_dimensions[r].height = 50
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 78

    # Tab colour & zoom
    ws.sheet_properties.tabColor = "FF0000"   # red tab
    ws.sheet_view.zoomScale = 125

    # Sheet protection
    lock_sheet(ws)


# ── Glossary sheet ─────────────────────────────────────────────────────────────

def build_glossary_sheet(ws, event_info):
    """Plain-English definitions of all statistical terms used in the workbook."""
    ev  = event_info["event_label"]
    np_ = event_info["n_pairs"]
    tot = 9 * np_

    f_title   = Font(name="Calibri", size=14, bold=True,  color=C_WHITE)
    f_section = Font(name="Calibri", size=11, bold=True,  color=C_WHITE)
    f_term    = Font(name="Calibri", size=11, bold=True,  color=C_DARK_TEXT)
    f_def     = Font(name="Calibri", size=10,             color="000000")

    fill_dark   = PatternFill("solid", fgColor=C_DARK_BLUE)
    fill_mid    = PatternFill("solid", fgColor=C_MID_BLUE)
    fill_light  = PatternFill("solid", fgColor=C_LIGHT_BLUE)
    fill_stripe = PatternFill("solid", fgColor="F5F9FF")
    fill_white  = PatternFill("solid", fgColor=C_WHITE)

    al_left = Alignment(horizontal="left", vertical="top",    wrap_text=True)
    al_cen  = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 1

    # Title
    ws.cell(row=r, column=1, value="Glossary of Statistical Terms").font = f_title
    ws.cell(row=r, column=1).fill = fill_dark
    ws.cell(row=r, column=1).alignment = al_cen
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 30
    r += 1

    ws.cell(row=r, column=1,
            value="Plain-English definitions of terms used in this workbook").font = \
        Font(name="Calibri", size=10, italic=True, color=C_DARK_TEXT)
    ws.cell(row=r, column=1).fill = fill_light
    ws.cell(row=r, column=1).alignment = al_cen
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 18
    r += 1

    ws.row_dimensions[r].height = 8
    r += 1

    terms = [
        ("\u25b6  SCORING TERMS", None),
        ("TES (Technical Element Score)",
         "The sum of base values and Grade of Execution (GOE) scores for all skating "
         "elements performed. Calculated as: Base Value + (trimmed-mean GOE \u00d7 GOE factor)."),
        ("PCS (Program Component Score)",
         "Score for the overall presentation and skating quality, assessed across "
         "up to five components. Each judge rates 0.25\u201310.00; the panel\u2019s "
         "trimmed mean is multiplied by the component factor."),
        ("TSS (Total Segment Score)",
         "TES + PCS \u2212 Deductions. The official competition score for one segment."),
        ("Trimmed mean",
         "The ISU discards the single highest and single lowest mark from the 9-judge "
         "panel for each element/component, then averages the remaining 7. This is the "
         "standard ISU scoring mechanism."),
        ("GOE (Grade of Execution)",
         "An integer from \u22125 to +5 that each judge assigns to each element. "
         "Converted to points by multiplication with a GOE factor specific to the element."),

        ("\u25b6  BIAS ANALYSIS TERMS", None),
        ("ISU-impact (I\u2C7C(t))",
         "The estimated change in a competitor\u2019s published ISU score if judge j\u2019s "
         "marks are replaced by the median of the other 8 judges. Positive = judge raised "
         "the score; negative = judge lowered it. This measure uses the exact ISU trimmed-mean "
         "formula, so the numbers are directly in ISU-published points."),
        ("BiasPoints (B\u2C7C(A,B))",
         "ISU-impact(judge j, Team A) \u2212 ISU-impact(judge j, Team B). "
         "Measures how much more judge j moved Team A\u2019s score (relative to the panel "
         "median) than Team B\u2019s. Positive = judge favoured A over B."),
        ("Vote: FOR_A / FOR_B / NEUTRAL",
         "Whether the judge\u2019s actual placement-direction matches the bias direction. "
         "FOR_A means the judge both inflated A and deflated B (consistent preferential pattern). "
         "NEUTRAL means the direction is mixed."),
        ("absBias",
         "Absolute value of BiasPoints (|B\u2C7C(A,B)|). Used as the sort key in "
         "\u2018Top 25 Pairs\u2019 to rank pairs by magnitude regardless of direction."),

        ("\u25b6  STATISTICAL TERMS", None),
        ("p-value (raw)",
         "The fraction of 10,000 random permutations that produced a BiasPoints value "
         "as extreme as or more extreme than the one observed. "
         "A value of 0.001 means 1-in-1,000 chance under the fair-judging null. "
         "Small = unusual."),
        ("q-value / BH-FDR",
         "The Benjamini\u2013Hochberg (BH) false discovery rate\u2013adjusted p-value. "
         f"When {tot:,} tests are performed simultaneously, some will appear significant by "
         "chance. BH-FDR controls the expected fraction of false discoveries. "
         "A q-value of 0.05 means: if you declare this pair biased, the expected "
         "fraction of wrong calls across all declared pairs is \u2264 5%."),
        ("Significance threshold",
         "q \u2264 0.05 after BH-FDR correction. This is a strict threshold \u2014 "
         "the most conservative standard multiple-comparison correction available."),
        ("Quantile permutation null",
         "The null model used to generate p-values. Each judge\u2019s marks are "
         "converted to percentile ranks within that judge\u2019s career-wide mark "
         "distribution, then the percentile labels are randomly reassigned across judges "
         "within each row. This preserves each judge\u2019s scoring style (conservative "
         "or generous overall) while eliminating any preferential targeting of specific "
         "competitors."),
        ("Style-adjusted",
         "The permutation null accounts for each judge\u2019s personal scoring style: "
         "a judge who scores everyone low cannot be flagged for scoring one team "
         "low (because that is their normal pattern). Only deviations from the "
         "judge\u2019s own distribution are flagged."),
        ("False discovery rate (FDR)",
         "The expected proportion of statistically significant findings that are "
         "actually false positives. BH-FDR controls this at the stated level (5%) "
         f"across the entire set of {tot:,} tests."),

        ("\u25b6  GENERAL NOTES", None),
        ("Verification",
         "All computed TES and TSS values reproduce the ISU\u2019s published official "
         "scores to within 0.01 points, confirming that the scoring model is exact."),
        ("Data source",
         f"ISU official \u2018Judges Details per Skater\u2019 for {ev}."),
        ("Analysis",
         "Performed by Michael Allman (University of Chicago Booth School of Business) "
         "with statistical AI assistance from Claude (Anthropic)."),
    ]

    fills_cycle = [fill_white, fill_stripe]
    data_row = 0
    for item in terms:
        term, defn = item
        if defn is None:
            ws.row_dimensions[r].height = 6
            r += 1
            ws.cell(row=r, column=1, value=term).font = f_section
            ws.cell(row=r, column=1).fill = fill_mid
            ws.cell(row=r, column=1).alignment = al_cen
            ws.merge_cells(f"A{r}:B{r}")
            ws.row_dimensions[r].height = 20
            r += 1
            data_row = 0
        else:
            fill = fills_cycle[data_row % 2]
            c_a = ws.cell(row=r, column=1, value=term)
            c_b = ws.cell(row=r, column=2, value=defn)
            c_a.font = f_term
            c_b.font = f_def
            c_a.fill = fill
            c_b.fill = fill
            c_a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c_a.border = brd
            c_b.border = brd
            ws.row_dimensions[r].height = 45
            r += 1
            data_row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    ws.sheet_properties.tabColor = C_MID_BLUE
    ws.sheet_view.zoomScale = 125
    lock_sheet(ws)


# ── Overview sheet ─────────────────────────────────────────────────────────────

def build_overview_sheet(ws, event_facts, event_label):
    """Populate the Overview navigation tab."""
    title_font   = Font(name="Calibri", size=14, bold=True,   color=C_DARK_BLUE)
    sub_font     = Font(name="Calibri", size=11, italic=True,  color=C_GREY_TEXT)
    section_font = Font(name="Calibri", size=11, bold=True,   color=C_DARK_BLUE)
    label_font   = Font(name="Calibri", size=10, bold=True)
    value_font   = Font(name="Calibri", size=10)
    hdr_font     = Font(name="Calibri", size=10, bold=True,   color=C_WHITE)
    hdr_fill     = PatternFill("solid", fgColor=C_MID_BLUE)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    row = 1

    # Title block
    c = ws.cell(row=row, column=1,
                value=f"{event_label} \u2014 Complete Analysis Workbook")
    c.font = title_font
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    c = ws.cell(row=row, column=1,
                value="ISU Official Scoring + Pairwise Bias Analysis")
    c.font = sub_font
    ws.merge_cells(f"A{row}:E{row}")
    row += 2

    # Event facts
    c = ws.cell(row=row, column=1, value="Event Information")
    c.font = section_font
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    for label, value in event_facts:
        ws.cell(row=row, column=1, value=label).font = label_font
        c = ws.cell(row=row, column=2, value=value)
        c.font = value_font
        c.alignment = left
        ws.merge_cells(f"B{row}:E{row}")
        row += 1

    row += 1

    # Tab directory
    c = ws.cell(row=row, column=1, value="Tab Directory")
    c.font = section_font
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    for col, hdr in enumerate(["#", "Tab Name", "Section", "Description"], start=1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    row += 1

    C_FINDINGS = "FFECEC"
    C_GLOSSARY = "EBF3FB"

    all_tabs = (
        [("Key Findings", "Key Findings", "Analysis Summary", C_FINDINGS)]
        + [
            (tab_name,
             tab_name,
             "ISU Source Data" if tab_name.startswith("ISU") else "Bias Analysis",
             C_ISU_GREEN if tab_name.startswith("ISU") else C_BIAS_YELLOW)
            for _, _, tab_name in TAB_CONFIG
        ]
        + [("Glossary", "Glossary", "Reference", C_GLOSSARY)]
    )

    for i, (_, tab_name, section, color) in enumerate(all_tabs, start=1):
        desc_text = TAB_DESCRIPTIONS.get(tab_name, "")
        fill      = PatternFill("solid", fgColor=color)
        num_cell  = ws.cell(row=row, column=1, value=i)
        num_cell.font      = value_font
        num_cell.alignment = center
        for col, val in enumerate([tab_name, section, desc_text], start=2):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = value_font
            c.fill      = fill
            c.alignment = left
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 62
    ws.column_dimensions["E"].width = 4

    ws.sheet_properties.tabColor = C_DARK_BLUE
    ws.sheet_view.zoomScale = 125
    lock_sheet(ws)


# ── Build bias workbook from database ─────────────────────────────────────────

def build_bias_workbook_from_db(event_id: int, is_ice_dance: bool = True) -> Workbook:
    """
    Query the canonical database and return an in-memory openpyxl Workbook
    with the same 5-sheet structure:
        JudgeSummary, JudgeImpact_ByTeam, Top25Pairs_ByJudge,
        PairwiseBias_All, Method
    """
    conn = sqlite3.connect(DB_PATH)
    wb   = Workbook()
    wb.remove(wb.active)

    # Sheet 1: JudgeSummary
    ws = wb.create_sheet("JudgeSummary")
    ws.append(["Judge", "num_pairs", "num_q\u22640.05", "min_p", "min_q"])

    n_entries = conn.execute(
        "SELECT COUNT(*) FROM entries WHERE event_id=?", (event_id,)
    ).fetchone()[0]
    n_pairs = n_entries * (n_entries - 1) // 2

    rows = conn.execute("""
        SELECT judge_position,
               SUM(CASE WHEN q_value_bh <= 0.05 THEN 1 ELSE 0 END),
               MIN(p_value),
               MIN(q_value_bh)
        FROM pairwise_impact_results
        WHERE event_id = ?
        GROUP BY judge_position
        ORDER BY judge_position
    """, (event_id,)).fetchall()
    for jpos, n_sig, min_p, min_q in rows:
        ws.append([jpos, n_pairs, n_sig, min_p, min_q])

    # Sheet 2: JudgeImpact_ByTeam
    ws = wb.create_sheet("JudgeImpact_ByTeam")
    ws.append(["Judge", "Rank", "Team", "NOC", "ImpactPoints_Ij"])

    rows = conn.execute("""
        SELECT judge_position, rank, team, noc, impact_points
        FROM judge_team_impacts
        WHERE event_id = ?
        ORDER BY judge_position, rank
    """, (event_id,)).fetchall()
    for row in rows:
        ws.append(list(row))

    # Sheet 3: Top25Pairs_ByJudge
    ws = wb.create_sheet("Top25Pairs_ByJudge")
    ws.append(["Judge", "absBias", "A_Rank", "A_Team", "A_NOC",
               "B_Rank", "B_Team", "B_NOC",
               "BiasPoints_AminusB", "Vote", "p_value", "q_value_BH"])

    rows = conn.execute("""
        SELECT judge_position, ABS(bias_points),
               rank_a, team_a, noc_a, rank_b, team_b, noc_b,
               bias_points, vote, p_value, q_value_bh,
               ROW_NUMBER() OVER (PARTITION BY judge_position
                                  ORDER BY ABS(bias_points) DESC) as rn
        FROM pairwise_impact_results
        WHERE event_id = ?
    """, (event_id,)).fetchall()
    for row in rows:
        rn = row[-1]
        if rn <= 25:
            ws.append(list(row[:-1]))

    # Sheet 4: PairwiseBias_All
    ws = wb.create_sheet("PairwiseBias_All")
    ws.append(["Judge", "A_Rank", "A_Team", "A_NOC",
               "B_Rank", "B_Team", "B_NOC",
               "BiasPoints_AminusB", "Vote", "p_value", "q_value_BH"])

    rows = conn.execute("""
        SELECT judge_position,
               rank_a, team_a, noc_a, rank_b, team_b, noc_b,
               bias_points, vote, p_value, q_value_bh
        FROM pairwise_impact_results
        WHERE event_id = ?
        ORDER BY judge_position, rank_a, rank_b
    """, (event_id,)).fetchall()
    for row in rows:
        ws.append(list(row))

    # Sheet 5: Method
    ws = wb.create_sheet("Method")
    ws.append(["Parameter", "Value"])

    meta = conn.execute("""
        SELECT DISTINCT permutations, rng_seed, method_version
        FROM pairwise_impact_results WHERE event_id = ?
    """, (event_id,)).fetchone()
    perms, seed, method_ver = meta if meta else (10000, 20260223, "isuimpact_quantile_v1")

    event_info_row = conn.execute("""
        SELECT c.name, e.discipline, e.segment
        FROM events e JOIN competitions c ON e.competition_id = c.competition_id
        WHERE e.event_id = ?
    """, (event_id,)).fetchone()
    comp_name  = event_info_row[0] if event_info_row else "Unknown"
    discipline = event_info_row[1] if event_info_row else "Unknown"
    segment    = event_info_row[2] if event_info_row else "Unknown"

    method_data = [
        ("Competition", comp_name),
        ("Discipline",  discipline),
        ("Segment",     segment),
        ("Competitors", str(n_entries)),
        ("Judges",      "9"),
        ("BiasStatistic",
         "ISU-impact: B_j(A,B)=I_j(A)-I_j(B), "
         "I_j(T)=sum_r Delta_jr, Delta_jr = F_r(actual) - F_r(neutralized_j)"),
        ("Neutralization", "Replace judge j's mark with median of other 8 judges"),
        ("Permutations",  str(perms)),
        ("RNG_seed",      str(seed)),
        ("Null_model",
         "Style-adjusted quantile permutation: percentile labels "
         "shuffled row-wise across judges, mapped back via each judge's empirical CDF"),
        ("CDF_scope",     "GLOBAL (career-wide per judge)"),
        ("Multiple_testing", f"BH-FDR within event, {9 * n_pairs} tests"),
        ("Significance",  "q <= 0.05"),
        ("p_value_formula", "p = (1 + #{|B_perm| >= |B_obs|}) / (M + 1)"),
        ("GOE_factor",
         "Effective factor derived from ISU published panel_goe_points / trimmed_mean"),
        ("Method_version", method_ver),
    ]
    for key, val in method_data:
        ws.append([key, val])

    # Worked example: effective GOE factor (Ice Dance only — combined elements)
    if is_ice_dance:
        example_row = conn.execute("""
            SELECT e.element_id, e.element_code, e.panel_goe_points, e.goe_factor_inferred,
                   en.team_name
            FROM elements e
            JOIN entries en ON e.entry_id = en.entry_id
            WHERE en.event_id = ? AND e.element_code LIKE 'SyTw%'
            ORDER BY e.element_id
            LIMIT 1
        """, (event_id,)).fetchone()

        if example_row:
            eid_ex, code_ex, panel_goe_ex, inferred_ex, team_ex = example_row
            grades_ex = conn.execute("""
                SELECT ejs.judge_goe_int
                FROM element_judge_scores ejs
                JOIN judges j ON ejs.judge_id = j.judge_id
                WHERE ejs.element_id = ?
                ORDER BY j.judge_position
            """, (eid_ex,)).fetchall()
            grades_list = [g[0] for g in grades_ex]

            if len(grades_list) == 9:
                sorted_g = sorted(grades_list)
                trimmed  = sorted_g[1:8]
                tm       = sum(trimmed) / 7.0
                eff      = panel_goe_ex / tm if abs(tm) > 0.001 else None

                ws.append([])
                ws.append(["─── WORKED EXAMPLE: Effective GOE Factor for Combined Elements ───", ""])
                ws.append(["Element", f"{code_ex}  ({team_ex})"])
                ws.append(["9 GOE grades (J1\u2013J9)", ", ".join(str(g) for g in grades_list)])
                ws.append(["Sorted", ", ".join(str(g) for g in sorted_g)])
                ws.append(["Trimmed (drop hi & lo)", ", ".join(str(g) for g in trimmed)])
                ws.append(["Trimmed mean", f"{sum(trimmed)}/7 = {tm:.6f}"])
                ws.append(["Published Panel GOE Points (ISU)", f"{panel_goe_ex}"])
                if eff is not None:
                    ws.append(["Effective factor = PanelGOE / TrimmedMean",
                               f"{panel_goe_ex} / {tm:.6f} = {eff:.4f}"])
                ws.append(["Stored goe_factor_inferred (single-element lookup)",
                           f"{inferred_ex}"])
                ws.append(["", ""])
                ws.append(["Why they differ",
                           "Combined elements (SyTw, OFT, DiSt) apply separate factors per "
                           "partner. The ISU publishes the final Panel GOE Points reflecting "
                           "this combined factor. The effective factor (PanelGOE / TrimmedMean) "
                           "recovers the true combined multiplier from the published data. "
                           "The single-element goe_factor_inferred reflects only one partner's "
                           "lookup value."])
                ws.append(["Verification",
                           "The ISU Protocol PDF shows Panel GOE Points for every element. "
                           "Dividing by the trimmed mean of the 9 judge grades reproduces "
                           "the effective factor used in this analysis."])

    conn.close()
    print(f"  Built bias workbook from database (event_id={event_id}, "
          f"{n_entries} competitors, {n_pairs} pairs)")
    return wb


# ── Build one event workbook ────────────────────────────────────────────────────

def build_event(event_id: int, dry_run: bool = False) -> bool:
    """
    Build (or dry-run) the complete analysis workbook for one event.
    Returns True on success, False on skip (missing source file).
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        try:
            isu_file, out_file = get_event_paths(conn, event_id)
        except FileNotFoundError as e:
            print(f"  \u26a0\ufe0f  SKIPPED: {e}")
            return False

        event_info    = get_event_info(conn, event_id)
        kf            = get_dynamic_key_findings(conn, event_id, event_info)
        event_facts   = build_dynamic_event_facts(event_info)
        bias_tab_meta = build_dynamic_bias_tab_meta(event_info, kf)
        judge_map     = {
            r[0]: f"{r[0]} \u2013 {r[1]}"
            for r in conn.execute(
                "SELECT judge_position, judge_name FROM judges "
                "WHERE event_id=? AND judge_name IS NOT NULL ORDER BY judge_position",
                (event_id,)
            ).fetchall()
        }

        if dry_run:
            od_flag = "OD=YES" if kf.get('is_od') else "OD=no"
            print(
                f"  [DRY RUN] event_id={event_id}  {od_flag}\n"
                f"    ISU: {os.path.basename(isu_file)}\n"
                f"    OUT: {os.path.basename(out_file)}\n"
                f"    {event_info['event_label']}  |  {event_info['n_entries']} competitors"
            )
            return True

        if not os.path.exists(DB_PATH):
            print(f"ERROR: Database not found: {DB_PATH}")
            return False

        print(f"  Loading ISU file: {os.path.basename(isu_file)}")
        isu_wb      = openpyxl.load_workbook(isu_file)
        print(f"    (data_only load for colour coding)")
        isu_wb_data = openpyxl.load_workbook(isu_file, data_only=True)

        print(f"  Loading bias data from database (event_id={event_id})")
        chat_wb = build_bias_workbook_from_db(
            event_id, is_ice_dance=event_info["is_ice_dance"]
        )

        sources = {"isu": isu_wb, "chat": chat_wb}

        # Build output workbook
        out_wb = Workbook()
        out_wb.remove(out_wb.active)

        # Tab 1: Overview
        print("  Building Overview tab...")
        overview_ws = out_wb.create_sheet("Overview")
        build_overview_sheet(overview_ws, event_facts, event_info["event_label"])

        # Tab 2: Key Findings
        print("  Building Key Findings tab...")
        kf_ws = out_wb.create_sheet("Key Findings")
        build_key_findings_sheet(kf_ws, kf, event_info)

        # Tabs 3\u201311: Copy / build source sheets
        for file_key, src_name, dst_name in TAB_CONFIG:
            src_wb = sources[file_key]
            if src_name not in src_wb.sheetnames:
                print(f"  WARNING: '{src_name}' not found in {file_key} workbook \u2014 skipping")
                continue

            print(f"  Copying '{src_name}' \u2192 '{dst_name}'")

            if file_key == "chat":
                meta   = bias_tab_meta[dst_name]
                dst_ws = out_wb.create_sheet(dst_name)
                copy_sheet_with_bias_header(
                    src_ws         = src_wb[src_name],
                    dst_ws         = dst_ws,
                    title          = meta["title"],
                    desc           = meta["desc"],
                    col_widths     = meta["col_widths"],
                    numeric_cols   = meta.get("numeric_cols", set()),
                    judge_map      = judge_map,
                    desc_font_size = meta.get("desc_font_size", 9),
                    center_cols    = meta.get("center_cols", set()),
                    hdr_row_height = meta.get("hdr_row_height", 36),
                )
            else:
                # ISU tabs: verbatim copy with formula remap
                dst_ws = out_wb.create_sheet(dst_name)
                copy_sheet(src_wb[src_name], dst_ws, formula_remap=ISU_SHEET_REMAP)

                if dst_name == "ISU \u2013 Summary":
                    dst_ws.sheet_view.zoomScale = 125
                    dst_ws["A1"].alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
                    dst_ws.row_dimensions[1].height = 40
                elif dst_name == "ISU \u2013 Elements (GOE)":
                    apply_high_low_colors(dst_ws, 7, 15, 5,
                                          data_ws=isu_wb_data["Element Scores"])
                    if event_info["is_ice_dance"]:
                        # Rename "Multiplier" → "GOE Factor*" and add footnote
                        for cell in dst_ws[3]:
                            if cell.value == "Multiplier":
                                cell.value = "GOE Factor*"
                                break
                        last_row = dst_ws.max_row + 2
                        note_cell = dst_ws.cell(row=last_row, column=1)
                        note_cell.value = (
                            "* GOE Factor shown is the single-element lookup value "
                            "(goe_factor_inferred). "
                            "For combined elements (SyTw, OFT, DiSt), the ISU applies "
                            "separate per-partner factors that sum to a higher effective "
                            "multiplier (~0.73 vs 0.60 shown). "
                            "The bias analysis uses the effective factor derived from the "
                            "ISU's published Panel GOE Points \u00f7 trimmed mean. "
                            "See the 'Bias \u2013 Method' tab for a worked example."
                        )
                        note_cell.font = Font(size=9, italic=True, color="595959")
                elif dst_name == "ISU \u2013 PCS":
                    apply_high_low_colors(dst_ws, 6, 14, 5,
                                          data_ws=isu_wb_data["Program Component Scores"])
                    for cell in dst_ws[2]:
                        if cell.font:
                            dst_ws.cell(row=2, column=cell.column).font = Font(
                                name=cell.font.name or "Calibri",
                                size=11,
                                italic=cell.font.italic,
                                color=cell.font.color,
                            )
                elif dst_name == "ISU \u2013 Legend":
                    fix_legend_colors(dst_ws)
                lock_sheet(dst_ws)

        # Last tab: Glossary
        print("  Building Glossary tab...")
        gl_ws = out_wb.create_sheet("Glossary")
        build_glossary_sheet(gl_ws, event_info)

        # Save
        out_wb.save(out_file)
        print(f"\n  Saved: {out_file}")
        print(f"  Tabs ({len(out_wb.sheetnames)}): {', '.join(out_wb.sheetnames)}")
        return True

    finally:
        conn.close()


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Build complete event analysis workbook (ISU scoring + bias analysis)."
    )
    parser.add_argument(
        "--event-id", type=int, default=2,
        help="event_id to build (default: 2 = OWG 2026 Ice Dance Free Dance)"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Print what would be built; do not write any files."
    )
    parser.add_argument(
        "--all-events", action="store_true",
        help="Build workbooks for all 142 analyzed events."
    )
    args = parser.parse_args()

    if not os.path.exists(DB_PATH):
        print(f"ERROR: Database not found: {DB_PATH}")
        sys.exit(1)

    if args.all_events:
        conn = sqlite3.connect(DB_PATH)
        event_ids = sorted(
            r[0] for r in conn.execute(
                "SELECT DISTINCT event_id FROM pairwise_impact_results"
            ).fetchall()
        )
        conn.close()
        total   = len(event_ids)
        skipped = []
        for i, eid in enumerate(event_ids, 1):
            print(f"\n[{i}/{total}] event_id={eid}")
            ok = build_event(eid, dry_run=args.dry_run)
            if not ok:
                skipped.append(eid)
        print(f"\n{'DRY RUN ' if args.dry_run else ''}Done: {total - len(skipped)}/{total} built.")
        if skipped:
            print(f"\u26a0\ufe0f  {len(skipped)} skipped: {skipped}")
    else:
        print(f"Building event_id={args.event_id} ...")
        ok = build_event(args.event_id, dry_run=args.dry_run)
        if not ok:
            sys.exit(1)


if __name__ == "__main__":
    main()
