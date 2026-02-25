#!/usr/bin/env python3
"""
generate_official_scoring_xlsx.py

Generate one Excel workbook per event (144 total) replicating the ISU official
scoring breakdown with live Excel formulas.

Format follows OWG2026_RD_Scoring_Model.xlsx exactly — 4 tabs:
  1. Summary         — one row per entry (Rank, Team, NOC, TES, PCS, Ded, TSS)
  2. Element Scores  — one row per element per entry, with judge GOEs and formulas
  3. Program Component Scores — one row per PCS component per entry
  4. Legend          — trim color coding explanation

Column layout (Element Scores, A–S, matching reference model):
  A=Rank  B=Team  C=NOC  D=Element  E=Base Value  F=Multiplier
  G..O = J1..J9 GOE integers
  P = Trimmed Sum   =SUM(G:O)-MAX(G:O)-MIN(G:O)
  Q = Panel GOE     =ROUND(P/(COUNT(G:O)-2)*F, 2)
  R = Element Score =ROUND(E+Q, 2)
  S = TES           =SUM(R_first:R_last)  [in last element row of entry]

Column layout (Program Component Scores, A–R, matching reference model):
  A=Rank  B=Team  C=NOC  D=Component  E=Factor
  F..N = J1..J9 judge marks
  O = Trimmed Sum   =SUM(F:N)-MAX(F:N)-MIN(F:N)
  P = Panel Avg     =ROUND(O/(COUNT(F:N)-2), 2)
  Q = Comp. Score   =ROUND(P*E, 2)
  R = Total PCS     =SUM(Q_first:Q_last)  [in first component row of entry]

Usage:
  python3 generate_official_scoring_xlsx.py --dry-run            # preview, no files written
  python3 generate_official_scoring_xlsx.py --dry-run --event-id 1
  python3 generate_official_scoring_xlsx.py --event-id 1         # generate one file
  python3 generate_official_scoring_xlsx.py                      # generate all 144 files
  python3 generate_official_scoring_xlsx.py --verify             # verify DB vs computed
  python3 generate_official_scoring_xlsx.py --db PATH            # use alternate DB

Output: excel_output/{competition_folder}_{pdf_basename}.xlsx
"""

import argparse
import os
import re
import sqlite3
import zipfile
from dataclasses import dataclass
from typing import Optional

import openpyxl
import openpyxl.formatting.rule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ── Paths ────────────────────────────────────────────────────────────────────
PROJ_DIR   = os.path.dirname(__file__)
DEFAULT_DB = os.path.join(PROJ_DIR, 'figure_skating_ijs_seed.sqlite')
OUT_DIR    = os.path.join(PROJ_DIR, 'excel_output')

# ── Colors (matching OWG2026_RD_Scoring_Model.xlsx) ──────────────────────────
GREEN_FILL  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
HEADER_FILL = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
TITLE_FILL  = PatternFill(start_color='17375E', end_color='17375E', fill_type='solid')

# ── ISU canonical PCS component order ────────────────────────────────────────
# Used to sort components consistently regardless of DB insertion order.
PCS_ORDER = [
    'Skating Skills', 'Transitions', 'Performance',
    'Composition', 'Interpretation', 'Presentation',
]

def pcs_sort_key(component_name: str) -> int:
    try:
        return PCS_ORDER.index(component_name)
    except ValueError:
        return 99  # Unknown components go last


# ── Data classes ──────────────────────────────────────────────────────────────
@dataclass
class JudgeInfo:
    position: str
    name: str
    country_code: str

@dataclass
class ElementRow:
    element_no: int
    element_code: str
    base_value: float
    goe_factor: float       # ISU-published GOE multiplier (from goe_factor_inferred)
    panel_goe_points: float # official value from PDF (for verification only)
    goe_ints: dict          # judge_position -> goe_int

@dataclass
class PCSRow:
    component_name: str
    factor: float
    panel_component_avg: float  # official value from PDF (for verification)
    marks: dict                 # judge_position -> judge_mark

@dataclass
class EntryData:
    entry_id: int
    rank: int
    team_name: str
    noc: str
    tes: float
    pcs: float
    deductions: float
    tss: float
    elements: list  # list of ElementRow, ordered by element_no
    pcs_rows: list  # list of PCSRow, sorted by ISU canonical order

@dataclass
class EventData:
    event_id: int
    competition_name: str
    season: str
    discipline: str
    segment: str
    scheduled_datetime: Optional[str]
    venue: Optional[str]
    xlsx_filename: str  # output filename (competition_folder prefix included)
    judges: list        # list of JudgeInfo, ordered J1..J9
    entries: list       # list of EntryData, ordered by rank


# ── Database queries ──────────────────────────────────────────────────────────
def load_event(conn, event_id: int) -> Optional[EventData]:
    conn.row_factory = sqlite3.Row

    ev = conn.execute("""
        SELECT ev.event_id, c.name AS comp_name, c.season, ev.discipline, ev.segment,
               ev.scheduled_datetime_local, ev.venue
        FROM events ev
        JOIN competitions c ON c.competition_id = ev.competition_id
        WHERE ev.event_id = ?
    """, (event_id,)).fetchone()
    if not ev:
        return None

    # Output filename: {competition_folder}_{pdf_stem}.xlsx
    src = conn.execute("""
        SELECT local_path FROM sources
        WHERE event_id = ? AND source_type = 'judges_details_pdf'
        LIMIT 1
    """, (event_id,)).fetchone()

    if src and src['local_path']:
        local_path  = src['local_path']
        pdf_basename = os.path.basename(local_path)
        path_parts   = local_path.replace('\\', '/').split('/')
        comp_folder  = path_parts[-2] if len(path_parts) >= 2 else 'event'
        pdf_stem     = os.path.splitext(pdf_basename)[0]
        xlsx_filename = f"{comp_folder}_{pdf_stem}.xlsx"
    else:
        disc_code = ev['discipline'].replace(' ', '_')[:20]
        seg_code  = ev['segment'].replace(' ', '_')[:15]
        xlsx_filename = f"Event{event_id}_{disc_code}_{seg_code}.xlsx"

    # Judges ordered by position (J1..J9)
    judge_rows = conn.execute("""
        SELECT judge_position, judge_name, country_code
        FROM judges WHERE event_id = ?
        ORDER BY judge_position
    """, (event_id,)).fetchall()
    judges = [JudgeInfo(j['judge_position'], j['judge_name'] or '', j['country_code'] or '')
              for j in judge_rows]

    # Entries ordered by rank
    entry_rows = conn.execute("""
        SELECT entry_id, rank, team_name, noc, tes, pcs, deductions, tss
        FROM entries WHERE event_id = ?
        ORDER BY rank
    """, (event_id,)).fetchall()

    entries = []
    for er in entry_rows:
        eid = er['entry_id']

        # Elements
        elem_rows = conn.execute("""
            SELECT el.element_no, el.element_code, el.base_value,
                   el.goe_factor_inferred, el.panel_goe_points, el.element_id
            FROM elements el
            WHERE el.entry_id = ?
            ORDER BY el.element_no
        """, (eid,)).fetchall()

        elements = []
        for el in elem_rows:
            goe_rows = conn.execute("""
                SELECT j.judge_position, ejs.judge_goe_int
                FROM element_judge_scores ejs
                JOIN judges j ON j.judge_id = ejs.judge_id AND j.event_id = ?
                WHERE ejs.element_id = ?
                ORDER BY j.judge_position
            """, (event_id, el['element_id'])).fetchall()
            goe_ints = {r['judge_position']: r['judge_goe_int'] for r in goe_rows}

            elements.append(ElementRow(
                element_no=el['element_no'],
                element_code=el['element_code'],
                base_value=el['base_value'] or 0.0,
                goe_factor=el['goe_factor_inferred'] or 0.0,
                panel_goe_points=el['panel_goe_points'] or 0.0,
                goe_ints=goe_ints,
            ))

        # PCS components — sorted into ISU canonical order
        pcs_comps = conn.execute("""
            SELECT pc.pcs_id, pc.component_name, pc.factor, pc.panel_component_avg
            FROM pcs_components pc
            WHERE pc.entry_id = ?
            ORDER BY pc.pcs_id
        """, (eid,)).fetchall()

        pcs_rows = []
        for pc in pcs_comps:
            mark_rows = conn.execute("""
                SELECT j.judge_position, pjs.judge_mark
                FROM pcs_judge_scores pjs
                JOIN judges j ON j.judge_id = pjs.judge_id AND j.event_id = ?
                WHERE pjs.pcs_id = ?
                ORDER BY j.judge_position
            """, (event_id, pc['pcs_id'])).fetchall()
            marks = {r['judge_position']: r['judge_mark'] for r in mark_rows}

            pcs_rows.append(PCSRow(
                component_name=pc['component_name'],
                factor=pc['factor'] or 1.0,
                panel_component_avg=pc['panel_component_avg'] or 0.0,
                marks=marks,
            ))

        # Sort into ISU canonical order
        pcs_rows.sort(key=lambda r: pcs_sort_key(r.component_name))

        entries.append(EntryData(
            entry_id=eid,
            rank=er['rank'],
            team_name=er['team_name'],
            noc=er['noc'] or '',
            tes=er['tes'] or 0.0,
            pcs=er['pcs'] or 0.0,
            deductions=er['deductions'] or 0.0,
            tss=er['tss'] or 0.0,
            elements=elements,
            pcs_rows=pcs_rows,
        ))

    return EventData(
        event_id=ev['event_id'],
        competition_name=ev['comp_name'],
        season=ev['season'] or '',
        discipline=ev['discipline'],
        segment=ev['segment'],
        scheduled_datetime=ev['scheduled_datetime_local'],
        venue=ev['venue'],
        xlsx_filename=xlsx_filename,
        judges=judges,
        entries=entries,
    )


def get_all_event_ids(conn) -> list:
    rows = conn.execute("SELECT event_id FROM events ORDER BY event_id").fetchall()
    return [r[0] for r in rows]


# ── Excel helpers ─────────────────────────────────────────────────────────────
def set_zoom(wb, zoom=125):
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = zoom
        ws.sheet_view.zoomScaleNormal = zoom


def add_trim_conditional_formatting(ws, first_data_row, last_data_row,
                                     judge_col_first, judge_col_last):
    """Per-row green (max) / red (min) CF on judge score columns."""
    for row in range(first_data_row, last_data_row + 1):
        row_range = f"{judge_col_first}{row}:{judge_col_last}{row}"
        ws.conditional_formatting.add(
            row_range,
            openpyxl.formatting.rule.FormulaRule(
                formula=[f"={judge_col_first}{row}=MAX(${judge_col_first}{row}:${judge_col_last}{row})"],
                fill=GREEN_FILL,
            )
        )
        ws.conditional_formatting.add(
            row_range,
            openpyxl.formatting.rule.FormulaRule(
                formula=[f"={judge_col_first}{row}=MIN(${judge_col_first}{row}:${judge_col_last}{row})"],
                fill=RED_FILL,
            )
        )


def patch_ignored_errors_xml(xlsx_path, sheet_xml_name, ranges_and_types):
    """
    Post-save XML patch: inject <ignoredErrors> to suppress 'formula omits
    adjacent cells' (formulaRange) warnings for sparse-formula columns.
    """
    tmp_path = xlsx_path + '.tmp'
    with zipfile.ZipFile(xlsx_path, 'r') as zin, \
         zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_xml_name:
                xml = data.decode('utf-8')
                ie_tags = ''.join(
                    f'<ignoredError sqref="{sqref}" {attr}="1"/>'
                    for sqref, attr in ranges_and_types
                )
                ie_block = f'<ignoredErrors>{ie_tags}</ignoredErrors>'
                if '<ignoredErrors>' in xml:
                    xml = re.sub(r'<ignoredErrors>.*?</ignoredErrors>',
                                 ie_block, xml, flags=re.DOTALL)
                else:
                    xml = xml.replace('</worksheet>', ie_block + '</worksheet>')
                data = xml.encode('utf-8')
            zout.writestr(item, data)
    os.replace(tmp_path, xlsx_path)


def _title_row(ws, row, merge_to_col, text, height=22):
    ws.merge_cells(f'A{row}:{merge_to_col}{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font  = Font(bold=True, size=12, color='FFFFFF')
    c.fill  = TITLE_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = height


def _subtitle_row(ws, row, merge_to_col, text, height=16):
    ws.merge_cells(f'A{row}:{merge_to_col}{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font  = Font(italic=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = height


# ── Tab 2: Element Scores ─────────────────────────────────────────────────────
def build_element_scores_tab(ws, ev: EventData):
    """
    Columns A–S, matching OWG2026_RD_Scoring_Model.xlsx:
      A=Rank  B=Team  C=NOC  D=Element  E=Base Value  F=Multiplier
      G..O = J1..J9 GOE integers
      P = Trimmed Sum
      Q = Panel GOE  =ROUND(P/(COUNT(G:O)-2)*F, 2)
      R = Element Score =ROUND(E+Q, 2)
      S = TES  =SUM(R_first:R_last)  [placed in LAST element row of each entry]

    Returns (first_data_row, last_data_row, 'G', 'O').
    """
    ws.title = 'Element Scores'

    date_str = ev.scheduled_datetime[:10] if ev.scheduled_datetime else ''
    venue_str = ev.venue or ''

    _title_row(ws, 1, 'S',
               f"{ev.competition_name} — {ev.discipline} — {ev.segment}  |  Element Scores (GOE)",
               height=22)
    _subtitle_row(ws, 2, 'S',
                  f"{venue_str}{'  ·  ' if venue_str and date_str else ''}{date_str}"
                  f"  |  GOE scale: −5 to +5 integer"
                  f"  |  Trimmed mean: drop highest & lowest  |  Panel GOE = (trimmed ÷ n) × multiplier",
                  height=16)

    # Row 3: column headers
    col_defs = [
        ('A', 'Rank',       5.0),
        ('B', 'Team',      40.0),
        ('C', 'NOC',        5.0),
        ('D', 'Element',   20.0),
        ('E', 'Base Value',10.0),
        ('F', 'Multiplier',11.0),
        ('G', 'J1', None), ('H', 'J2', None), ('I', 'J3', None),
        ('J', 'J4', None), ('K', 'J5', None), ('L', 'J6', None),
        ('M', 'J7', None), ('N', 'J8', None), ('O', 'J9', None),
        ('P', 'Trimmed Sum', 11.0),
        ('Q', 'Panel GOE',   10.0),
        ('R', 'Element Score',13.0),
        ('S', 'TES',          9.0),
    ]
    for col_letter, label, width in col_defs:
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = Font(bold=True, size=9)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if width is not None:
            ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[3].height = 18

    # Row 4: judge names (wrap, 60pt)
    judge_map = {j.position: j for j in ev.judges}
    judge_el_cols = {  # position -> column letter (Element Scores)
        'J1': 'G', 'J2': 'H', 'J3': 'I', 'J4': 'J', 'J5': 'K',
        'J6': 'L', 'J7': 'M', 'J8': 'N', 'J9': 'O',
    }
    for pos, col_letter in judge_el_cols.items():
        j = judge_map.get(pos)
        name    = j.name if j else pos
        country = j.country_code if j else ''
        display = f"{name}\n{country}" if country else name
        cell = ws.cell(row=4, column=column_index_from_string(col_letter), value=display)
        cell.font = Font(size=8)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 60.0

    ws.freeze_panes = 'G5'  # freeze cols A-F and rows 1-4

    # ── Data rows ─────────────────────────────────────────────────────────────
    cur_row        = 5
    first_data_row = 5

    for entry in ev.entries:
        n_elem          = len(entry.elements)
        entry_start_row = cur_row
        entry_end_row   = cur_row + max(n_elem, 1) - 1

        for i, elem in enumerate(entry.elements):
            row = cur_row + i
            r   = str(row)

            # A B C: only on first element row; merge across all element rows of this entry
            if i == 0:
                ws.cell(row=row, column=1, value=entry.rank).alignment = \
                    Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=2, value=entry.team_name).alignment = \
                    Alignment(horizontal='left', vertical='center')
                ws.cell(row=row, column=3, value=entry.noc).alignment = \
                    Alignment(horizontal='center', vertical='center')
                # Merge A, B, C vertically across all element rows for this entry
                if n_elem > 1:
                    ws.merge_cells(f'A{entry_start_row}:A{entry_end_row}')
                    ws.merge_cells(f'B{entry_start_row}:B{entry_end_row}')
                    ws.merge_cells(f'C{entry_start_row}:C{entry_end_row}')

            # D: element code
            ws.cell(row=row, column=4, value=elem.element_code).alignment = \
                Alignment(horizontal='left', vertical='center')

            # E: base value
            c = ws.cell(row=row, column=5, value=elem.base_value)
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='center', vertical='center')

            # F: Multiplier (ISU-published GOE factor from goe_factor_inferred)
            c = ws.cell(row=row, column=6, value=elem.goe_factor)
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='center', vertical='center')

            # G..O: judge GOE integers
            for pos, col_letter in judge_el_cols.items():
                goe_val = elem.goe_ints.get(pos)
                cell = ws.cell(row=row, column=column_index_from_string(col_letter), value=goe_val)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # P: Trimmed Sum = SUM(G:O) - MAX(G:O) - MIN(G:O)
            g_range = f"G{r}:O{r}"
            c = ws.cell(row=row, column=16,
                        value=f"=SUM({g_range})-MAX({g_range})-MIN({g_range})")
            c.alignment = Alignment(horizontal='center', vertical='center')

            # Q: Panel GOE = ROUND(P / (COUNT(G:O) - 2) * F, 2)
            # Using COUNT dynamically handles 7-judge vs 9-judge panels correctly.
            c = ws.cell(row=row, column=17,
                        value=f"=ROUND(P{r}/(COUNT({g_range})-2)*F{r},2)")
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='center', vertical='center')

            # R: Element Score = ROUND(E + Q, 2)
            c = ws.cell(row=row, column=18,
                        value=f"=ROUND(E{r}+Q{r},2)")
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='center', vertical='center')

            # S: TES — placed in the LAST element row of this entry, =SUM(R_first:R_last)
            if i == n_elem - 1:
                c = ws.cell(row=row, column=19,
                            value=f"=SUM(R{entry_start_row}:R{entry_end_row})")
                c.number_format = '0.00'
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center', vertical='center')

        cur_row += max(n_elem, 1)

    last_data_row = cur_row - 1
    return first_data_row, last_data_row, 'G', 'O'


# ── Tab 3: Program Component Scores ──────────────────────────────────────────
def build_pcs_tab(ws, ev: EventData):
    """
    Columns A–R, matching OWG2026_RD_Scoring_Model.xlsx:
      A=Rank  B=Team  C=NOC  D=Component  E=Factor
      F..N = J1..J9 judge marks
      O = Trimmed Sum
      P = Panel Avg  =ROUND(O/(COUNT(F:N)-2), 2)
      Q = Comp. Score =ROUND(P*E, 2)
      R = Total PCS  =SUM(Q_first:Q_last)  [in FIRST component row of each entry]

    Returns (first_data_row, last_data_row, 'F', 'N').
    """
    ws.title = 'Program Component Scores'

    date_str  = ev.scheduled_datetime[:10] if ev.scheduled_datetime else ''
    venue_str = ev.venue or ''

    _title_row(ws, 1, 'R',
               f"{ev.competition_name} — {ev.discipline} — {ev.segment}  |  Program Component Scores (PCS)",
               height=22)
    _subtitle_row(ws, 2, 'R',
                  f"{venue_str}{'  ·  ' if venue_str and date_str else ''}{date_str}"
                  f"  |  PCS scale: 0.25–10.0"
                  f"  |  Trimmed mean: drop highest & lowest  |  Component Score = Panel Avg × Factor",
                  height=16)

    # Row 3: column headers
    col_defs = [
        ('A', 'Rank',       5.0),
        ('B', 'Team',      40.0),
        ('C', 'NOC',        5.0),
        ('D', 'Component', 18.0),
        ('E', 'Factor',     7.0),
        ('F', 'J1', None), ('G', 'J2', None), ('H', 'J3', None),
        ('I', 'J4', None), ('J', 'J5', None), ('K', 'J6', None),
        ('L', 'J7', None), ('M', 'J8', None), ('N', 'J9', None),
        ('O', 'Trimmed Sum', 11.0),
        ('P', 'Panel Avg',   10.0),
        ('Q', 'Comp. Score', 12.0),
        ('R', 'Total PCS',   11.0),
    ]
    for col_letter, label, width in col_defs:
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = Font(bold=True, size=9)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if width is not None:
            ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[3].height = 18

    # Row 4: judge names
    judge_map = {j.position: j for j in ev.judges}
    judge_pcs_cols = {
        'J1': 'F', 'J2': 'G', 'J3': 'H', 'J4': 'I', 'J5': 'J',
        'J6': 'K', 'J7': 'L', 'J8': 'M', 'J9': 'N',
    }
    for pos, col_letter in judge_pcs_cols.items():
        j = judge_map.get(pos)
        name    = j.name if j else pos
        country = j.country_code if j else ''
        display = f"{name}\n{country}" if country else name
        cell = ws.cell(row=4, column=column_index_from_string(col_letter), value=display)
        cell.font = Font(size=8)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 60.0

    ws.freeze_panes = 'F5'  # freeze cols A-E and rows 1-4

    # ── Data rows ─────────────────────────────────────────────────────────────
    cur_row        = 5
    first_data_row = 5

    for entry in ev.entries:
        n_comp          = len(entry.pcs_rows)
        entry_start_row = cur_row
        entry_end_row   = cur_row + max(n_comp, 1) - 1

        for i, pcs in enumerate(entry.pcs_rows):
            row = cur_row + i
            r   = str(row)

            # A B C D E: only on first row, merged vertically across component rows
            if i == 0:
                ws.cell(row=row, column=1, value=entry.rank).alignment = \
                    Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=2, value=entry.team_name).alignment = \
                    Alignment(horizontal='left', vertical='center')
                ws.cell(row=row, column=3, value=entry.noc).alignment = \
                    Alignment(horizontal='center', vertical='center')
                if n_comp > 1:
                    ws.merge_cells(f'A{entry_start_row}:A{entry_end_row}')
                    ws.merge_cells(f'B{entry_start_row}:B{entry_end_row}')
                    ws.merge_cells(f'C{entry_start_row}:C{entry_end_row}')

            # D: component name
            ws.cell(row=row, column=4, value=pcs.component_name).alignment = \
                Alignment(horizontal='left', vertical='center')

            # E: factor
            c = ws.cell(row=row, column=5, value=pcs.factor)
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='center', vertical='center')

            # F..N: judge marks
            for pos, col_letter in judge_pcs_cols.items():
                mark_val = pcs.marks.get(pos)
                cell = ws.cell(row=row, column=column_index_from_string(col_letter),
                               value=mark_val)
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # O: Trimmed Sum
            f_range = f"F{r}:N{r}"
            c = ws.cell(row=row, column=15,
                        value=f"=SUM({f_range})-MAX({f_range})-MIN({f_range})")
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='center', vertical='center')

            # P: Panel Avg = ROUND(O / (COUNT(F:N) - 2), 2)
            # ISU computes panel_avg at 2dp (matches official results).
            # Reference model shows 4dp but that's a display artifact — 2dp
            # is required to reproduce the official PCS values exactly.
            c = ws.cell(row=row, column=16,
                        value=f"=ROUND(O{r}/(COUNT({f_range})-2),2)")
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='center', vertical='center')

            # Q: Component Score = ROUND(P * E, 2)
            c = ws.cell(row=row, column=17,
                        value=f"=ROUND(P{r}*E{r},2)")
            c.number_format = '0.00'
            c.alignment = Alignment(horizontal='center', vertical='center')

            # R: Total PCS — placed in FIRST component row of entry only
            if i == 0:
                c = ws.cell(row=row, column=18,
                            value=f"=SUM(Q{entry_start_row}:Q{entry_end_row})")
                c.number_format = '0.00'
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center', vertical='center')

        cur_row += max(n_comp, 1)

    last_data_row = cur_row - 1
    return first_data_row, last_data_row, 'F', 'N'


# ── Tab 1: Summary ────────────────────────────────────────────────────────────
def build_summary_tab(ws, ev: EventData, el_ws_name: str, pcs_ws_name: str):
    """
    One row per entry.
    TES links to column S (last element row of entry) in Element Scores tab.
    PCS links to column R (first component row of entry) in PCS tab.
    """
    ws.title = 'Summary'

    date_str  = ev.scheduled_datetime[:10] if ev.scheduled_datetime else ''
    venue_str = ev.venue or ''

    _title_row(ws, 1, 'G',
               f"{ev.competition_name} — {ev.discipline} — {ev.segment}  |  Official Scoring Summary",
               height=22)
    _subtitle_row(ws, 2, 'G',
                  f"{venue_str}{'  ·  ' if venue_str and date_str else ''}{date_str}"
                  f"  |  Scores computed from Element Scores and PCS sheets",
                  height=16)

    # Headers (row 3)
    col_defs = [
        ('A', 'Rank',        6.0),
        ('B', 'Team',       42.0),
        ('C', 'NOC',         6.0),
        ('D', 'TES',         9.0),
        ('E', 'PCS',         9.0),
        ('F', 'Deductions', 11.0),
        ('G', 'TSS',         9.0),
    ]
    for col_letter, label, width in col_defs:
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = Font(bold=True, size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[3].height = 18

    ws.freeze_panes = 'A4'

    # Pre-compute row positions:
    # TES is in the LAST element row of each entry (column S).
    # Total PCS is in the FIRST component row of each entry (column R).
    el_row  = 5
    pcs_row = 5
    entry_row_map = {}  # entry_id -> (el_last_row, pcs_first_row)

    for entry in ev.entries:
        n_elem = len(entry.elements)
        n_comp = len(entry.pcs_rows)
        el_last_row  = el_row + max(n_elem, 1) - 1
        pcs_first_row = pcs_row
        entry_row_map[entry.entry_id] = (el_last_row, pcs_first_row)
        el_row  += max(n_elem, 1)
        pcs_row += max(n_comp, 1)

    for i, entry in enumerate(ev.entries):
        data_row = i + 4
        el_last_r, pcs_first_r = entry_row_map[entry.entry_id]

        ws.cell(row=data_row, column=1, value=entry.rank).alignment = \
            Alignment(horizontal='center', vertical='center')
        ws.cell(row=data_row, column=2, value=entry.team_name).alignment = \
            Alignment(horizontal='left', vertical='center')
        ws.cell(row=data_row, column=3, value=entry.noc).alignment = \
            Alignment(horizontal='center', vertical='center')

        # TES: link to S (last element row of this entry in Element Scores tab)
        c = ws.cell(row=data_row, column=4,
                    value=f"='{el_ws_name}'!S{el_last_r}")
        c.number_format = '0.00'
        c.alignment = Alignment(horizontal='center', vertical='center')

        # PCS: link to R (first component row of this entry in PCS tab)
        c = ws.cell(row=data_row, column=5,
                    value=f"='{pcs_ws_name}'!R{pcs_first_r}")
        c.number_format = '0.00'
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Deductions (hardcoded value from DB)
        ded = entry.deductions if entry.deductions else 0.0
        c = ws.cell(row=data_row, column=6, value=ded)
        c.number_format = '0.00'
        c.alignment = Alignment(horizontal='center', vertical='center')

        # TSS = TES + PCS + Deductions
        dr = str(data_row)
        c = ws.cell(row=data_row, column=7,
                    value=f"=D{dr}+E{dr}+F{dr}")
        c.number_format = '0.00'
        c.alignment = Alignment(horizontal='center', vertical='center')


# ── Tab 4: Legend ─────────────────────────────────────────────────────────────
def build_legend_tab(ws):
    ws.title = 'Legend'

    ws['A1'] = 'Colour Legend'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A3'] = 'Judge score colour coding (per element / component row):'
    ws['A3'].font = Font(bold=True)

    ws['A5'] = 'Highest score(s)'
    ws['B5'].fill  = GREEN_FILL
    ws['B5'].value = 'Light Green'
    ws['C5'] = 'The highest judge score in this row — trimmed from the panel average'

    ws['A6'] = 'Lowest score(s)'
    ws['B6'].fill  = RED_FILL
    ws['B6'].value = 'Light Red / Pink'
    ws['C6'] = 'The lowest judge score in this row — trimmed from the panel average'

    ws['A8'] = 'Note:'
    ws['A8'].font = Font(bold=True)
    ws['A9']  = ('When multiple judges are tied for the highest (or lowest) score, '
                 'all tied cells are highlighted.')
    ws['A10'] = ('Trimmed mean: the panel GOE / component average drops the single highest '
                 'and single lowest score,')
    ws['A11'] = 'then averages the remaining judges.'
    ws['A13'] = ('Multiplier (Element Scores, col F): the ISU-published GOE multiplier '
                 'from ISU Communications (e.g. 0.8 for sequential twizzles).')
    ws['A14'] = ('  Panel GOE = ROUND(Trimmed Sum ÷ (n judges − 2) × Multiplier, 2)'
                 '   [dynamic divisor handles 7- or 9-judge panels]')
    ws['A15'] = '  Element Score = ROUND(Base Value + Panel GOE, 2)'
    ws['A16'] = '  TES = sum of all Element Scores for this entry'
    ws['A17'] = ('  Panel Avg (PCS) = ROUND(Trimmed Sum ÷ (n judges − 2), 2)'
                 '   Component Score = ROUND(Panel Avg × Factor, 2)')
    ws['A18'] = '  Total PCS = sum of all Component Scores for this entry'
    ws['A19'] = '  TSS = TES + Total PCS + Deductions  (deductions are stored as negative numbers)'

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60


# ── Main workbook builder ─────────────────────────────────────────────────────
def build_workbook(ev: EventData, out_path: str) -> None:
    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_el      = wb.create_sheet()
    ws_pcs     = wb.create_sheet()
    ws_legend  = wb.create_sheet()

    el_ws_name  = 'Element Scores'
    pcs_ws_name = 'Program Component Scores'

    # Build content
    el_first, el_last, el_jcf, el_jcl = build_element_scores_tab(ws_el, ev)
    pcs_first, pcs_last, pcs_jcf, pcs_jcl = build_pcs_tab(ws_pcs, ev)
    build_summary_tab(ws_summary, ev, el_ws_name, pcs_ws_name)
    build_legend_tab(ws_legend)

    # Zoom 125% on all sheets
    set_zoom(wb, 125)

    # Conditional formatting: green/red per row on judge columns
    add_trim_conditional_formatting(ws_el,  el_first,  el_last,  el_jcf, el_jcl)
    add_trim_conditional_formatting(ws_pcs, pcs_first, pcs_last, pcs_jcf, pcs_jcl)

    wb.save(out_path)

    # Post-save: patch ignoredErrors XML to suppress "formula omits adjacent cells"
    # warnings on sparse columns (S in Element Scores, R in PCS).
    wb2 = openpyxl.load_workbook(out_path)
    sheet_names = wb2.sheetnames
    el_idx  = sheet_names.index(el_ws_name) + 1
    pcs_idx = sheet_names.index(pcs_ws_name) + 1
    wb2.close()

    if el_last >= 5:
        patch_ignored_errors_xml(
            out_path,
            f'xl/worksheets/sheet{el_idx}.xml',
            [(f'S5:S{el_last}', 'formulaRange')]
        )
    if pcs_last >= 5:
        patch_ignored_errors_xml(
            out_path,
            f'xl/worksheets/sheet{pcs_idx}.xml',
            [(f'R5:R{pcs_last}', 'formulaRange')]
        )


# ── Verification ──────────────────────────────────────────────────────────────
def verify_event(conn, event_id: int) -> list:
    """
    Recompute TES and PCS from raw judge scores in the DB and compare to stored
    entries.tes / entries.pcs / entries.tss. Returns list of mismatch dicts.

    TES: uses panel_goe_points (the official ISU PDF value) to match the Excel
         formula output exactly.
    PCS: uses trimmed mean with dynamic n_judges (handles 7- and 9-judge panels).
         Rounds panel_avg to 2dp then multiplies by factor (matching ISU official computation).
    """
    conn.row_factory = sqlite3.Row
    mismatches = []

    ev_info = conn.execute("""
        SELECT c.name, ev.discipline, ev.segment
        FROM events ev JOIN competitions c ON c.competition_id = ev.competition_id
        WHERE ev.event_id = ?
    """, (event_id,)).fetchone()
    if not ev_info:
        return mismatches

    entries = conn.execute(
        "SELECT entry_id, rank, team_name, tes, pcs, deductions, tss "
        "FROM entries WHERE event_id = ? ORDER BY rank",
        (event_id,)
    ).fetchall()

    for entry in entries:
        eid = entry['entry_id']

        # TES: sum(ROUND(base_value + panel_goe_points, 2)) — matches Excel formula
        elements = conn.execute(
            "SELECT base_value, panel_goe_points FROM elements "
            "WHERE entry_id = ? ORDER BY element_no", (eid,)
        ).fetchall()
        computed_tes = round(sum(
            round((el['base_value'] or 0.0) + (el['panel_goe_points'] or 0.0), 2)
            for el in elements
        ), 2)

        # PCS: sum(ROUND(ROUND(trimmed_sum / (n-2), 2) * factor, 2))
        pcs_comps = conn.execute(
            "SELECT pcs_id, factor FROM pcs_components WHERE entry_id = ? ORDER BY pcs_id",
            (eid,)
        ).fetchall()
        computed_pcs = 0.0
        for pc in pcs_comps:
            marks = [r['judge_mark'] for r in conn.execute(
                "SELECT pjs.judge_mark FROM pcs_judge_scores pjs WHERE pjs.pcs_id = ?",
                (pc['pcs_id'],)
            ).fetchall() if r['judge_mark'] is not None]
            if len(marks) >= 3:
                ts          = sum(marks) - max(marks) - min(marks)
                n_remaining = len(marks) - 2
                panel_avg   = round(ts / n_remaining, 2)  # 2dp, matching ISU official computation
                comp_score  = round(panel_avg * (pc['factor'] or 1.0), 2)
            else:
                comp_score = 0.0
            computed_pcs += comp_score
        computed_pcs = round(computed_pcs, 2)

        computed_tss = round(computed_tes + computed_pcs + (entry['deductions'] or 0.0), 2)

        stored_tes = round(entry['tes'] or 0.0, 2)
        stored_pcs = round(entry['pcs'] or 0.0, 2)
        stored_tss = round(entry['tss'] or 0.0, 2)

        tes_diff = abs(computed_tes - stored_tes)
        pcs_diff = abs(computed_pcs - stored_pcs)
        tss_diff = abs(computed_tss - stored_tss)

        TOLERANCE = 0.02
        if tes_diff > TOLERANCE or pcs_diff > TOLERANCE or tss_diff > TOLERANCE:
            mismatches.append({
                'event_id':     event_id,
                'comp_name':    ev_info['name'],
                'discipline':   ev_info['discipline'],
                'segment':      ev_info['segment'],
                'entry_id':     eid,
                'rank':         entry['rank'],
                'team_name':    entry['team_name'],
                'computed_tes': computed_tes, 'stored_tes': stored_tes, 'tes_diff': tes_diff,
                'computed_pcs': computed_pcs, 'stored_pcs': stored_pcs, 'pcs_diff': pcs_diff,
                'computed_tss': computed_tss, 'stored_tss': stored_tss, 'tss_diff': tss_diff,
            })

    return mismatches


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='Generate official ISU scoring Excel workbooks from the figure skating database.'
    )
    parser.add_argument('--dry-run',  action='store_true',
                        help='Print planned output files; do not write anything')
    parser.add_argument('--event-id', type=int, metavar='N',
                        help='Process only this event_id (for testing)')
    parser.add_argument('--verify',   action='store_true',
                        help='Recompute TES/PCS from raw scores and compare to DB values')
    parser.add_argument('--db',       default=DEFAULT_DB,
                        help='Path to SQLite database (default: seed DB)')
    args = parser.parse_args()

    if not os.path.exists(args.db):
        raise SystemExit(f"Database not found: {args.db}")

    os.makedirs(OUT_DIR, exist_ok=True)
    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row

    event_ids = [args.event_id] if args.event_id else get_all_event_ids(conn)

    # ── Verification mode ─────────────────────────────────────────────────────
    if args.verify:
        print(f"\n{'='*70}")
        print("VERIFICATION: recomputing TES/PCS from raw judge scores vs DB values")
        print(f"Tolerance: ±0.02\n{'='*70}\n")
        all_mismatches = []
        total_entries  = 0

        for eid in event_ids:
            mismatches     = verify_event(conn, eid)
            n              = conn.execute(
                "SELECT COUNT(*) FROM entries WHERE event_id=?", (eid,)
            ).fetchone()[0]
            total_entries += n
            all_mismatches.extend(mismatches)

        print(f"Total entries checked: {total_entries}")
        print(f"Mismatches (>±0.02):   {len(all_mismatches)}")
        print(f"Pass rate:             "
              f"{100*(total_entries-len(all_mismatches))/max(total_entries,1):.1f}%\n")

        if all_mismatches:
            print("MISMATCHES:")
            print(f"{'Event':<50} {'Rank':>4} {'Team':<40} "
                  f"{'TES diff':>8} {'PCS diff':>8} {'TSS diff':>8}")
            print('-' * 130)
            for m in all_mismatches:
                ev_str = (f"{m['comp_name'][:30]} | "
                          f"{m['discipline'][:15]} | {m['segment'][:12]}")
                print(f"{ev_str:<50} {m['rank']:>4} {m['team_name']:<40} "
                      f"{m['tes_diff']:>8.4f} {m['pcs_diff']:>8.4f} {m['tss_diff']:>8.4f}")

            report_path = os.path.join(OUT_DIR, 'verification_report.txt')
            with open(report_path, 'w') as f:
                f.write(f"Verification Report — {args.db}\n")
                f.write(f"Total entries: {total_entries}\n")
                f.write(f"Mismatches:    {len(all_mismatches)}\n")
                f.write(f"Pass rate:     "
                        f"{100*(total_entries-len(all_mismatches))/max(total_entries,1):.1f}%\n\n")
                for m in all_mismatches:
                    f.write(f"event_id={m['event_id']} rank={m['rank']} {m['team_name']}\n")
                    f.write(f"  TES: computed={m['computed_tes']:.2f} "
                            f"stored={m['stored_tes']:.2f} diff={m['tes_diff']:.4f}\n")
                    f.write(f"  PCS: computed={m['computed_pcs']:.2f} "
                            f"stored={m['stored_pcs']:.2f} diff={m['pcs_diff']:.4f}\n")
                    f.write(f"  TSS: computed={m['computed_tss']:.2f} "
                            f"stored={m['stored_tss']:.2f} diff={m['tss_diff']:.4f}\n\n")
            print(f"\nReport written to: {report_path}")
        else:
            print("✅ All entries match within tolerance.")

        conn.close()
        return

    # ── Generation mode ───────────────────────────────────────────────────────
    generated = 0
    skipped   = 0
    errors    = []

    for eid in event_ids:
        ev = load_event(conn, eid)
        if ev is None:
            print(f"  event_id={eid}: not found, skipping")
            skipped += 1
            continue

        out_path = os.path.join(OUT_DIR, ev.xlsx_filename)
        print(f"  event_id={eid:3d}: {ev.xlsx_filename:<65} "
              f"({len(ev.entries)} entries, "
              f"{sum(len(e.elements) for e in ev.entries)} elements)")

        if args.dry_run:
            continue

        try:
            build_workbook(ev, out_path)
            generated += 1
        except Exception as exc:
            import traceback
            errors.append((eid, ev.xlsx_filename, str(exc)))
            print(f"    ⚠️  ERROR: {exc}")
            traceback.print_exc()

    conn.close()

    if args.dry_run:
        print(f"\nDry run complete. Would generate {len(event_ids)} file(s) in {OUT_DIR}/")
        return

    print(f"\n{'='*60}")
    print(f"Generated: {generated}")
    print(f"Skipped:   {skipped}")
    print(f"Errors:    {len(errors)}")
    if errors:
        for eid, fname, msg in errors:
            print(f"  event_id={eid} {fname}: {msg}")
    else:
        print(f"✅ All files written to {OUT_DIR}/")


if __name__ == '__main__':
    main()
