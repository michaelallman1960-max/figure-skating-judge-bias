#!/usr/bin/env python3
"""
generate_event_report.py
========================
Generates a 7-tab OSNR Analysis Excel report for any figure skating event
stored in figure_skating_ijs_v4.sqlite.

Usage:
    python3 generate_event_report.py --event-id 1
    python3 generate_event_report.py --event-id 1 --output /path/to/output.xlsx

Tabs:
  1. Summary            - One-page overview with standings, panel, OSNR flags, LOJO
  2. Raw Scores         - Full raw judge GOE scores per element
  3. Regime Comparison  - R1 (official) vs R0 (raw mean) vs R2 (OSNR-adjusted)
  4. Judge Statistics   - Per-judge B(j), LOJO, significance counts
  5. Pairwise B(j)      - All pairs with p <= 0.10, autofilter
  6. LOJO Counterfactual- Heatmap of counterfactual ranks if each judge removed
  7. Legend             - Color coding and methodology notes
"""

import argparse
import os
import sqlite3
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Default paths
BASE_DIR = "/Users/allman/Library/CloudStorage/Dropbox/Dropbox Mike/Judging Bias"
DB_PATH  = os.path.join(BASE_DIR, "figure_skating_ijs_v4.sqlite")

def parse_args():
    p = argparse.ArgumentParser(description="Generate OSNR Event Report Excel file")
    p.add_argument("--event-id", type=int, required=True, help="Event ID in the database")
    p.add_argument("--output", type=str, default=None,
                   help="Output path (default: auto-generated from event info)")
    p.add_argument("--db", type=str, default=DB_PATH,
                   help=f"Path to SQLite database (default: {DB_PATH})")
    return p.parse_args()

def solid_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def std_border():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)

def sect_border_fn(col, section_boundaries):
    t = Side(style="thin")
    k = Side(style="medium")
    return Border(left=t, right=(k if col in section_boundaries else t), top=t, bottom=t)

FILL_DARK    = solid_fill("2F4F7F")
FILL_GRAY    = solid_fill("D3D3D3")
FILL_LGRAY   = solid_fill("F2F2F2")
FILL_WHITE   = solid_fill("FFFFFF")
FILL_BLUE    = solid_fill("BDD7EE")
FILL_GREEN   = solid_fill("C6EFCE")
FILL_YELLOW  = solid_fill("FFEB9C")
FILL_SALMON  = solid_fill("FCE4D6")
FILL_PALE_Y  = solid_fill("FFFFD0")

def mk_font(size=10, bold=False, italic=False, color="000000", name="Calibri"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

FONT_TITLE  = mk_font(13, bold=True, color="FFFFFF")
FONT_WHBOLD = mk_font(10, bold=True, color="FFFFFF")
FONT_BOLD   = mk_font(10, bold=True)
FONT_DATA   = mk_font(10)
FONT_JUDG   = mk_font(9, italic=True)

CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LTOP = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

GOE_COLORS = {
    -5:"FF0000",-4:"FF4444",-3:"FF8888",-2:"FFBBBB",-1:"FFDDDD",
     0:"FFFFFF", 1:"DDFFDD", 2:"BBFFBB", 3:"88FF88", 4:"44FF44", 5:"00CC00",
}

def goe_fill(goe_int):
    return PatternFill(fill_type="solid", fgColor=GOE_COLORS.get(int(goe_int),"FFFFFF"))

def trimmed_mean(ints, drop=1):
    s = sorted(ints)
    trimmed = s[drop:len(s)-drop]
    return sum(trimmed)/len(trimmed) if trimmed else 0.0

def raw_mean(values):
    return sum(values)/len(values) if values else 0.0

def rank_by_tss(tss_map):
    sorted_items = sorted(tss_map.items(), key=lambda kv: kv[1], reverse=True)
    return {eid: i for i,(eid,_) in enumerate(sorted_items,1)}

def delta_str_ranks(r1, r_alt):
    d = r_alt - r1
    if d == 0: return u"\u2014"
    elif d < 0: return f"+{abs(d)}"
    else: return f"-{d}"

def delta_color_ranks(r1, r_alt):
    d = r_alt - r1
    if d < 0: return "006400"
    if d > 0: return "CC0000"
    return "808080"

def load_event_data(conn, event_id):
    cur = conn.cursor()
    cur.execute("""
        SELECT e.event_id, e.discipline, e.segment, e.level,
               e.scheduled_datetime_local, e.venue,
               c.name, c.season, c.location
        FROM events e JOIN competitions c ON e.competition_id=c.competition_id
        WHERE e.event_id=?
    """, (event_id,))
    ev = cur.fetchone()
    if not ev:
        raise ValueError(f"No event found with event_id={event_id}")
    event_info = {
        "event_id":ev[0],"discipline":ev[1],"segment":ev[2],
        "level":ev[3],"datetime_local":ev[4],"venue":ev[5],
        "competition_name":ev[6],"season":ev[7],"location":ev[8],
    }
    cur.execute("SELECT label_key,label_value FROM event_labels WHERE event_id=?", (event_id,))
    event_info["labels"] = dict(cur.fetchall())

    cur.execute("""SELECT judge_id,judge_position,judge_name,country_code
        FROM judges WHERE event_id=? ORDER BY judge_position""", (event_id,))
    judges = {}
    judge_id_to_pos = {}
    for jid,jpos,jname,jcc in cur.fetchall():
        judges[jpos] = {"id":jid,"name":jname or "","country":jcc or ""}
        judge_id_to_pos[jid] = jpos
    judge_positions = sorted(judges.keys())

    # Tier 2 rule: a judge is flagged for removal only if they satisfy BOTH:
    #   (1) podium_changes > 0  (LOJO outcome-determinative: changes any top-3 position), AND
    #   (2) at least one pairwise comparison with p <= 0.001 (sig001 > 0)
    # A judge who is merely a "swing vote" in a close competition (LOJO only,
    # no statistical anomaly) is NOT flagged — that reflects competitive
    # sensitivity, not bias.
    cur.execute("""
        SELECT les.judge_position, les.judge_name, les.judge_id
        FROM lojo_event_summary les
        JOIN (
            SELECT judge_position, SUM(is_significant_001) AS sig001
            FROM pairwise_judge_statistics
            WHERE event_id = ?
            GROUP BY judge_position
        ) pjs ON les.judge_position = pjs.judge_position
        WHERE les.event_id = ?
          AND les.podium_changes > 0
          AND pjs.sig001 > 0
        ORDER BY les.kendall_tau_distance DESC
        LIMIT 1
    """, (event_id, event_id))
    flagged_row = cur.fetchone()
    flagged_judge_pos = flagged_row[0] if flagged_row else None
    flagged_judge_id  = flagged_row[2] if flagged_row else None
    event_info["flagged_judge_pos"] = flagged_judge_pos
    event_info["flagged_judge_id"]  = flagged_judge_id

    cur.execute("""SELECT entry_id,start_no,team_name,noc,rank,tes,pcs,deductions,tss
        FROM entries WHERE event_id=? ORDER BY rank""", (event_id,))
    entries = []
    for row in cur.fetchall():
        eid,sno,tname,noc,rank,tes,pcs,ded,tss = row
        entries.append({
            "entry_id":eid,"start_no":sno,"team_name":tname,
            "noc":noc or "","rank":rank,
            "tes_r1":round(tes,2),"pcs_r1":round(pcs,2),
            "deductions":ded or 0.0,"tss_r1":round(tss,2),
        })

    for entry in entries:
        eid = entry["entry_id"]
        cur.execute("""SELECT element_id,element_no,element_code,base_value,panel_goe_points
            FROM elements WHERE entry_id=? ORDER BY element_no""", (eid,))
        elems_raw = cur.fetchall()
        elements = []
        for elem_id,elem_no,elem_code,bv,official_goe in elems_raw:
            cur.execute("""SELECT ejs.judge_id,j.judge_position,ejs.judge_goe_int
                FROM element_judge_scores ejs
                JOIN judges j ON ejs.judge_id=j.judge_id AND j.event_id=?
                WHERE ejs.element_id=? ORDER BY j.judge_position""", (event_id,elem_id))
            goe_by_pos = {}
            for g_jid,g_jpos,g_int in cur.fetchall():
                goe_by_pos[g_jpos] = {"judge_id":g_jid,"goe":g_int}
            all_ints = [goe_by_pos[jp]["goe"] for jp in judge_positions if jp in goe_by_pos]
            r1_trimmed = trimmed_mean(all_ints, drop=1)
            goe_increment = (official_goe/r1_trimmed) if r1_trimmed != 0 else 0.0
            r0_goe = round(raw_mean(all_ints)*goe_increment, 2)
            r1_goe = round(official_goe, 2)
            if flagged_judge_id:
                r2_ints = [goe_by_pos[jp]["goe"] for jp in judge_positions
                           if jp in goe_by_pos and goe_by_pos[jp]["judge_id"] != flagged_judge_id]
                r2_mean = trimmed_mean(r2_ints,drop=1) if len(r2_ints)>=3 else raw_mean(r2_ints)
                r2_goe  = round(r2_mean*goe_increment, 2)
            else:
                r2_goe = r1_goe
            elements.append({
                "element_no":elem_no,"element_code":elem_code,
                "base_value":round(bv,2),"goe_by_pos":goe_by_pos,
                "panel_goe_official":round(official_goe,2),
                "r1_goe":r1_goe,"r1_elem":round(bv+r1_goe,2),
                "r0_goe":r0_goe,"r0_elem":round(bv+r0_goe,2),
                "r2_goe":r2_goe,"r2_elem":round(bv+r2_goe,2),
            })
        entry["elements"] = elements

    for entry in entries:
        eid = entry["entry_id"]
        cur.execute("""SELECT pc.pcs_id,pc.factor FROM pcs_components pc
            WHERE pc.entry_id=? ORDER BY pc.component_name""", (eid,))
        pcs_rows = cur.fetchall()
        total_r0 = total_r2 = 0.0
        for pcs_id,factor in pcs_rows:
            cur.execute("""SELECT pjs.judge_id,pjs.judge_mark FROM pcs_judge_scores pjs
                WHERE pjs.pcs_id=? ORDER BY pjs.judge_id""", (pcs_id,))
            mark_rows = cur.fetchall()
            marks_all = [m for _,m in mark_rows]
            total_r0 += raw_mean(marks_all)*factor
            if flagged_judge_id:
                marks_r2 = [m for jid,m in mark_rows if jid!=flagged_judge_id]
                r2_comp  = trimmed_mean(marks_r2,drop=1) if len(marks_r2)>=3 else raw_mean(marks_r2)
            else:
                r2_comp = trimmed_mean(marks_all,drop=1)
            total_r2 += r2_comp*factor
        entry["pcs_r0"] = round(total_r0,2)
        entry["pcs_r2"] = round(total_r2,2)

    for entry in entries:
        # deductions are stored as negative numbers (e.g., -1.00 for a 1-point deduction)
        # so we ADD ded to the sum (not subtract), which correctly reduces the total.
        ded = entry["deductions"]
        entry["tes_r0"] = round(sum(el["r0_elem"] for el in entry["elements"]),2)
        entry["tes_r2"] = round(sum(el["r2_elem"] for el in entry["elements"]),2)
        entry["tss_r0"] = round(entry["tes_r0"]+entry["pcs_r0"]+ded,2)
        entry["tss_r2"] = round(entry["tes_r2"]+entry["pcs_r2"]+ded,2)

    r0_ranks = rank_by_tss({e["entry_id"]:e["tss_r0"] for e in entries})
    r2_ranks = rank_by_tss({e["entry_id"]:e["tss_r2"] for e in entries})
    for entry in entries:
        entry["rank_r0"] = r0_ranks[entry["entry_id"]]
        entry["rank_r2"] = r2_ranks[entry["entry_id"]]

    cur.execute("""SELECT judge_id,judge_position,judge_name,n_entries,
               winner_changes,podium_changes,n_rank_inversions,total_pairs,
               kendall_tau_distance,actual_winner_name,cf_winner_name,actual_margin,cf_margin
        FROM lojo_event_summary WHERE event_id=? ORDER BY judge_position""", (event_id,))
    lojo_summary = {}
    for row in cur.fetchall():
        jid,jpos,jname,n_ent,wc,pc,nri,tp,ktd,aw,cfw,am,cfm = row
        lojo_summary[jpos] = {
            "judge_id":jid,"judge_name":jname,
            "winner_changes":wc,"podium_changes":pc,
            "n_rank_inversions":nri,"total_pairs":tp,
            "kendall_tau":ktd,"actual_winner":aw,"cf_winner":cfw,
            "actual_margin":am,"cf_margin":cfm,
        }

    cur.execute("""SELECT ls.judge_position,ls.entry_id,ls.cf_tss,ls.cf_rank,
               ls.official_rank,ls.rank_change
        FROM lojo_scores ls WHERE ls.event_id=? ORDER BY ls.judge_position,ls.official_rank""", (event_id,))
    lojo_scores = {}
    for row in cur.fetchall():
        jpos,eid,cf_tss,cf_rank,off_rank,rchg = row
        if jpos not in lojo_scores: lojo_scores[jpos] = {}
        lojo_scores[jpos][eid] = {"cf_tss":cf_tss,"cf_rank":cf_rank,"official_rank":off_rank,"rank_change":rchg}

    cur.execute("""SELECT judge_id,judge_position,judge_name,judge_country,
               mean_goe_deviation,bias_z_score,correlation_with_panel,outlier_count
        FROM judge_event_statistics WHERE event_id=? ORDER BY judge_position""", (event_id,))
    judge_stats = {}
    for row in cur.fetchall():
        jid,jpos,jname,jcc,mgd,bz,corr,oc = row
        judge_stats[jpos] = {"judge_id":jid,"name":jname,"country":jcc or "",
                             "mean_goe_dev":mgd,"bias_z":bz,"correlation":corr,"outlier_count":oc}

    cur.execute("""SELECT judge_position,judge_id,
               SUM(is_significant_01),SUM(is_significant_001),MAX(bias_statistic)
        FROM pairwise_judge_statistics WHERE event_id=? GROUP BY judge_id ORDER BY judge_position""", (event_id,))
    pairwise_agg = {}
    for row in cur.fetchall():
        jpos,jid,s01,s001,maxb = row
        pairwise_agg[jpos] = {"sig01":s01 or 0,"sig001":s001 or 0,"max_bias":maxb or 0}

    cur.execute("""SELECT p1.judge_position,p1.skater_a_name,p1.skater_b_name,p1.bias_statistic
        FROM pairwise_judge_statistics p1
        WHERE p1.event_id=? AND p1.bias_statistic=(
            SELECT MAX(p2.bias_statistic) FROM pairwise_judge_statistics p2
            WHERE p2.event_id=? AND p2.judge_id=p1.judge_id)
        ORDER BY p1.judge_position""", (event_id,event_id))
    max_bias_rows = {}
    for row in cur.fetchall():
        jpos,sa,sb,bs = row
        if jpos not in max_bias_rows:
            max_bias_rows[jpos] = {"skater_a":sa,"skater_b":sb,"bias":bs}

    cur.execute("""SELECT judge_position,judge_name,judge_country,
               skater_a_name,skater_a_country,skater_a_rank,
               skater_b_name,skater_b_country,skater_b_rank,
               bias_statistic,p_value,is_significant_01,is_significant_001,is_significant_bonferroni,
               mean_deviation_a,mean_deviation_b,differential,num_elements_a,num_elements_b
        FROM pairwise_judge_statistics WHERE event_id=? AND p_value<=0.10
        ORDER BY p_value ASC""", (event_id,))
    pairwise_rows = cur.fetchall()

    cur.execute("""SELECT COUNT(*),SUM(is_significant_01),SUM(is_significant_001),SUM(is_significant_bonferroni)
        FROM pairwise_judge_statistics WHERE event_id=?""", (event_id,))
    pw_totals = cur.fetchone()

    return {
        "event_info":event_info,"judges":judges,"judge_positions":judge_positions,
        "judge_id_to_pos":judge_id_to_pos,"entries":entries,
        "lojo_summary":lojo_summary,"lojo_scores":lojo_scores,
        "judge_stats":judge_stats,"pairwise_agg":pairwise_agg,
        "max_bias_rows":max_bias_rows,"pairwise_rows":pairwise_rows,"pw_totals":pw_totals,
    }


def build_tab_summary(ws, data):
    ev = data["event_info"]
    judges = data["judges"]
    j_positions = data["judge_positions"]
    entries = data["entries"]
    lojo_sum = data["lojo_summary"]
    pairwise_agg = data["pairwise_agg"]

    col_widths = {1:6,2:30,3:5,4:8,5:8,6:5,7:8,8:28,9:5,10:12,11:5,12:5,13:18,14:28,15:10,16:10}
    for cn,w in col_widths.items():
        ws.column_dimensions[get_column_letter(cn)].width = w

    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = "OWG 2026 Ice Dance \u2014 Rhythm Dance | OSNR Analysis Report"
    c.font = FONT_TITLE; c.fill = FILL_DARK; c.alignment = CTR
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:P2")
    c = ws["A2"]
    c.value = "Olympic Winter Games 2026 \u00b7 Milano Ice Skating Arena \u00b7 February 9, 2026"
    c.font = mk_font(10,italic=True); c.fill = FILL_GRAY; c.alignment = CTR
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8

    # Box 1: Official Standings header
    r = 4
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "OFFICIAL STANDINGS & REGIME COMPARISON"
    c.font = FONT_WHBOLD; c.fill = FILL_DARK; c.alignment = CTR
    ws.row_dimensions[r].height = 18

    r = 5
    for ci,h in enumerate(["Rank","Team","NOC","R1 TSS","R0 TSS"],1):
        cell = ws.cell(row=r,column=ci)
        cell.value=h; cell.font=FONT_BOLD; cell.fill=FILL_BLUE; cell.alignment=CTR; cell.border=std_border()
    ws.row_dimensions[r].height = 18

    r = 6
    for entry in entries:
        rk = entry["rank"]
        if rk==1: bg=solid_fill("FFF2CC")
        elif rk==2: bg=solid_fill("F2F2F2")
        elif rk==3: bg=solid_fill("FCE4D6")
        elif rk%2==0: bg=FILL_WHITE
        else: bg=FILL_LGRAY
        d0 = delta_str_ranks(entry["rank"],entry["rank_r0"])
        d2 = delta_str_ranks(entry["rank"],entry["rank_r2"])
        for ci,(val,fmt,algn) in enumerate([(rk,"0",CTR),(entry["team_name"],None,LFT),(entry["noc"],None,CTR),(entry["tss_r1"],"0.00",CTR),(entry["tss_r0"],"0.00",CTR)],1):
            cell=ws.cell(row=r,column=ci)
            cell.value=val; cell.fill=bg; cell.alignment=algn
            cell.font=FONT_BOLD if ci==1 else FONT_DATA
            if fmt: cell.number_format=fmt
            cell.border=std_border()
        ws.row_dimensions[r].height=15
        r+=1

    r+=1
    ws.merge_cells(f"A{r}:E{r}")
    flagged_pos = ev.get("flagged_judge_pos")
    flagged_name = judges.get(flagged_pos,{}).get("name","") if flagged_pos else ""
    flagged_country = judges.get(flagged_pos,{}).get("country","") if flagged_pos else ""
    c=ws[f"A{r}"]
    if flagged_pos:
        c.value=f"R2 = OSNR-adjusted: {flagged_pos} ({flagged_name}, {flagged_country}) removed"
    else:
        c.value="R2 = No judge meets Tier 2 criteria (winner change + p\u22640.001) — R2 = R1"
    c.font=mk_font(9,italic=True); c.fill=solid_fill("FFFFD0"); c.alignment=LFT
    box1_end = r

    # Box 2: Judge Panel
    br2 = 4
    ws.merge_cells(f"G{br2}:K{br2}")
    c=ws[f"G{br2}"]
    c.value="Panel of Judges \u2014 Rhythm Dance"
    c.font=FONT_WHBOLD; c.fill=FILL_DARK; c.alignment=CTR

    br2=5
    for ci,h in zip([7,8,9,10],["Pos","Name","NOC","OSNR Status"]):
        cell=ws.cell(row=br2,column=ci)
        cell.value=h; cell.font=FONT_BOLD; cell.fill=FILL_BLUE; cell.alignment=CTR; cell.border=std_border()

    br2=6
    for jpos in j_positions:
        j=judges[jpos]
        pa=pairwise_agg.get(jpos,{})
        ls=lojo_sum.get(jpos,{})
        wc=ls.get("winner_changes",0); pc=ls.get("podium_changes",0)
        sig01=pa.get("sig01",0); sig001=pa.get("sig001",0)
        osnr=("Tier 2" if sig001>0 and pc>0 else "Tier 1" if sig01>0 else "No flag")
        osnr_fill=(solid_fill("C6EFCE") if osnr=="No flag" else solid_fill("FFEB9C") if osnr=="Tier 1" else solid_fill("F4B942"))
        for ci,val in zip([7,8,9,10],[jpos,j["name"],j["country"] or "\u2014",osnr]):
            cell=ws.cell(row=br2,column=ci)
            cell.value=val
            cell.font=(mk_font(9,color="808080") if val=="No flag" else FONT_DATA)
            cell.fill=(osnr_fill if ci==10 else FILL_WHITE)
            cell.alignment=(LFT if ci==8 else CTR); cell.border=std_border()
        ws.row_dimensions[br2].height=15; br2+=1

    br2+=1
    wc_count=sum(1 for ls in lojo_sum.values() if ls.get("podium_changes",0)>0)
    ws.merge_cells(f"G{br2}:K{br2}")
    c=ws[f"G{br2}"]
    c.value=f"LOJO Podium Changes: {wc_count}"
    c.font=mk_font(9,bold=(wc_count>0),color=("CC0000" if wc_count>0 else "000000"))
    c.fill=(solid_fill("FFFFD0") if wc_count>0 else FILL_WHITE); c.alignment=LFT
    br2+=1

    tau_vals=[ls["kendall_tau"] for ls in lojo_sum.values() if ls.get("kendall_tau") is not None]
    if tau_vals:
        ws.merge_cells(f"G{br2}:K{br2}")
        c=ws[f"G{br2}"]
        c.value=f"Kendall \u03c4 range: {min(tau_vals):.3f} \u2013 {max(tau_vals):.3f} across panel"
        c.font=mk_font(9,italic=True); c.fill=FILL_WHITE; c.alignment=LFT
    box2_end=br2

    # Box 3: OSNR Summary
    br3=4
    ws.merge_cells(f"M{br3}:P{br3}")
    c=ws[f"M{br3}"]
    c.value="OSNR Analysis \u2014 Rhythm Dance"
    c.font=FONT_WHBOLD; c.fill=FILL_DARK; c.alignment=CTR

    br3=5
    tier1=sum(1 for pa in pairwise_agg.values() if pa.get("sig01",0)>0)
    tier2=sum(1 for jpos in j_positions if pairwise_agg.get(jpos,{}).get("sig001",0)>0 and lojo_sum.get(jpos,{}).get("podium_changes",0)>0)
    for txt,color,bg in [
        (f"Tier 1 flags (p \u2264 0.01): {tier1}","808080" if tier1==0 else "CC6600",FILL_WHITE),
        (f"Tier 2 flags (p \u2264 0.001 + outcome-determinative): {tier2}","808080" if tier2==0 else "CC0000",FILL_WHITE),
        ("Note: No judge flagged in this segment" if tier1==0 and tier2==0 else "","808080",FILL_WHITE),
    ]:
        if not txt: continue
        ws.merge_cells(f"M{br3}:P{br3}")
        c=ws[f"M{br3}"]
        c.value=txt; c.font=mk_font(9,color=color); c.alignment=LFT; c.fill=bg
        br3+=1

    br3+=1
    ws.merge_cells(f"M{br3}:P{br3}")
    c=ws[f"M{br3}"]
    c.value="LOJO COUNTERFACTUAL HIGHLIGHTS"
    c.font=mk_font(9,bold=True); c.fill=FILL_BLUE; c.alignment=CTR
    br3+=1

    for ci,h in zip([13,14,15,16],["Judge","Kendall \u03c4","Winner\nChange","CF Winner"]):
        cell=ws.cell(row=br3,column=ci)
        cell.value=h; cell.font=FONT_BOLD; cell.fill=FILL_BLUE; cell.alignment=CTR; cell.border=std_border()
    br3+=1

    for jpos in j_positions:
        ls=lojo_sum.get(jpos,{})
        wc_flag=ls.get("winner_changes",0); pc_flag=ls.get("podium_changes",0)
        tau=ls.get("kendall_tau",0) or 0
        cf_winner=ls.get("cf_winner","") or ""
        actual_w=ls.get("actual_winner","") or ""
        cf_show=cf_winner if cf_winner!=actual_w else "\u2014"
        row_fill=solid_fill("FFFFD0") if pc_flag>0 else FILL_WHITE
        for ci,val in zip([13,14,15,16],[jpos,tau,"YES" if pc_flag>0 else "No",cf_show]):
            cell=ws.cell(row=br3,column=ci)
            cell.value=val; cell.fill=row_fill; cell.alignment=CTR
            cell.font=mk_font(9,bold=(ci==15 and wc_flag>0),color=("CC0000" if ci==15 and wc_flag>0 else "000000"))
            if ci==14 and isinstance(val,float): cell.number_format="0.00000"
            cell.border=std_border()
        ws.row_dimensions[br3].height=15; br3+=1

    box3_end=br3
    footer_row=max(box1_end,box2_end,box3_end)+2
    ws.merge_cells(f"A{footer_row}:P{footer_row}")
    c=ws[f"A{footer_row}"]
    c.value="Full analysis: Raw Scores | Regime Comparison | Judge Statistics | Pairwise B(j) | LOJO Counterfactual"
    c.font=mk_font(9,italic=True); c.fill=FILL_GRAY; c.alignment=CTR


def build_tab_raw_scores(ws, data):
    judges=data["judges"]; j_positions=data["judge_positions"]; entries=data["entries"]
    n_j=len(j_positions)
    C_J1=6; C_JN=C_J1+n_j-1
    C_PGOE=C_JN+1; C_ESCORE=C_JN+2; C_TES=C_JN+3; C_PCS=C_JN+4; C_TSS=C_JN+5
    TOTAL=C_TSS

    cw={1:6,2:28,3:5,4:18,5:9}
    for c in range(C_J1,C_JN+1): cw[c]=16
    for c in range(C_PGOE,TOTAL+1): cw[c]=9
    for cn,w in cw.items(): ws.column_dimensions[get_column_letter(cn)].width=w

    ws.merge_cells(f"A1:{get_column_letter(TOTAL)}1")
    c=ws["A1"]; c.value="OWG 2026 Ice Dance \u2014 Rhythm Dance | Raw Judge GOE Scores"
    c.font=FONT_TITLE; c.fill=FILL_DARK; c.alignment=CTR; ws.row_dimensions[1].height=22

    ws.merge_cells(f"A2:{get_column_letter(TOTAL)}2")
    c=ws["A2"]
    c.value="Olympic Winter Games 2026 \u00b7 Milano Ice Skating Arena \u00b7 February 9, 2026 | GOE scale: \u22125 to +5 | Trimmed mean: drop highest & lowest of 9"
    c.font=mk_font(10,italic=True); c.fill=FILL_GRAY; c.alignment=CTR; ws.row_dimensions[2].height=22

    ws.row_dimensions[3].height=28
    for ci,h in [(1,"Rank"),(2,"Team"),(3,"NOC"),(4,"Element"),(5,"Base\nValue")]:
        c=ws.cell(row=3,column=ci); c.value=h; c.font=FONT_BOLD; c.fill=FILL_GRAY; c.alignment=CTR; c.border=std_border()
    for i,jpos in enumerate(j_positions):
        c=ws.cell(row=3,column=C_J1+i); c.value=jpos; c.font=FONT_BOLD; c.fill=FILL_BLUE; c.alignment=CTR; c.border=std_border()
    for ci,h in [(C_PGOE,"Panel\nGOE"),(C_ESCORE,"Element\nScore"),(C_TES,"TES"),(C_PCS,"PCS"),(C_TSS,"TSS")]:
        c=ws.cell(row=3,column=ci); c.value=h; c.font=FONT_BOLD; c.fill=FILL_GREEN; c.alignment=CTR; c.border=std_border()

    ws.merge_cells("A4:E4")
    c=ws["A4"]; c.value="Judge Names \u2192"; c.font=FONT_JUDG; c.fill=FILL_BLUE; c.alignment=CTR
    ws.row_dimensions[4].height=28
    for i,jpos in enumerate(j_positions):
        c=ws.cell(row=4,column=C_J1+i); c.value=judges[jpos]["name"]; c.font=FONT_JUDG; c.fill=FILL_BLUE; c.alignment=CTR; c.border=std_border()
    for ci in range(C_PGOE,TOTAL+1): ws.cell(row=4,column=ci).fill=FILL_WHITE

    ws.freeze_panes="F5"
    cur_row=5
    for entry in entries:
        n=len(entry["elements"]); r_start=cur_row; r_end=cur_row+n-1; rk=entry["rank"]
        alt=FILL_LGRAY if rk%2==1 else FILL_WHITE
        for ei,el in enumerate(entry["elements"]):
            r=cur_row+ei; ws.row_dimensions[r].height=15
            c=ws.cell(row=r,column=4,value=el["element_code"]); c.fill=alt; c.font=FONT_DATA; c.alignment=LFT; c.border=std_border()
            c=ws.cell(row=r,column=5,value=el["base_value"]); c.fill=alt; c.alignment=CTR; c.number_format="0.00"; c.border=std_border()
            for j_idx,jpos in enumerate(j_positions):
                ci=C_J1+j_idx
                if jpos in el["goe_by_pos"]:
                    gi=el["goe_by_pos"][jpos]["goe"]
                    c=ws.cell(row=r,column=ci,value=gi); c.fill=goe_fill(gi); c.font=FONT_DATA; c.alignment=CTR; c.border=std_border()
                else: ws.cell(row=r,column=ci).fill=FILL_GRAY
            c=ws.cell(row=r,column=C_PGOE,value=el["panel_goe_official"]); c.fill=FILL_GREEN; c.number_format="0.00"; c.alignment=CTR; c.border=std_border()
            escore=round(el["base_value"]+el["panel_goe_official"],2)
            c=ws.cell(row=r,column=C_ESCORE,value=escore); c.fill=FILL_GREEN; c.number_format="0.00"; c.alignment=CTR; c.border=std_border()
            for ci2 in [1,2,3]: ws.cell(row=r,column=ci2).fill=FILL_GRAY; ws.cell(row=r,column=ci2).border=std_border()
            for ci2 in [C_TES,C_PCS,C_TSS]: ws.cell(row=r,column=ci2).fill=FILL_GREEN; ws.cell(row=r,column=ci2).border=std_border()

        def mv(col,val,fill,fmt=None,bold=False,algn=CTR):
            if n>1: ws.merge_cells(start_row=r_start,start_column=col,end_row=r_end,end_column=col)
            c=ws.cell(row=r_start,column=col); c.value=val; c.fill=fill; c.alignment=algn
            c.font=FONT_BOLD if bold else FONT_DATA
            if fmt: c.number_format=fmt; c.border=std_border()
        mv(1,rk,FILL_GRAY,"0",bold=True); mv(2,entry["team_name"],FILL_GRAY,bold=True,algn=LFT); mv(3,entry["noc"],FILL_GRAY)
        mv(C_TES,entry["tes_r1"],FILL_GREEN,"0.00",bold=True); mv(C_PCS,entry["pcs_r1"],FILL_GREEN,"0.00",bold=True); mv(C_TSS,entry["tss_r1"],FILL_GREEN,"0.00",bold=True)
        cur_row+=n


def build_tab_regime_comparison(ws, data):
    judges=data["judges"]; j_positions=data["judge_positions"]; entries=data["entries"]; ev=data["event_info"]
    flagged_pos=ev.get("flagged_judge_pos")
    flagged_name=judges.get(flagged_pos,{}).get("name","") if flagged_pos else ""
    flagged_cc=judges.get(flagged_pos,{}).get("country","") if flagged_pos else ""
    n_j=len(j_positions)
    CR=1;CT=2;CN=3;CE=4;CB=5;CJ1=6;CJ9=CJ1+n_j-1
    CR1G=CJ9+1;CR1E=CJ9+2;CR1T=CJ9+3;CR1P=CJ9+4;CR1S=CJ9+5
    CR0G=CR1S+1;CR0E=CR1S+2;CR0T=CR1S+3;CR0P=CR1S+4;CR0S=CR1S+5
    CR2G=CR0S+1;CR2E=CR0S+2;CR2T=CR0S+3;CR2P=CR0S+4;CR2S=CR0S+5
    TOTAL=CR2S
    SB=[CB,CJ9,CR1S,CR0S,CR2S]

    def sb(col): return sect_border_fn(col,SB)

    cw={1:6,2:28,3:5,4:18,5:9}
    for c in range(6,CJ9+1): cw[c]=16
    for c in range(CR1G,TOTAL+1): cw[c]=9
    for cn,w in cw.items(): ws.column_dimensions[get_column_letter(cn)].width=w
    for r,h in [(1,22),(2,30),(3,22),(4,28),(5,36)]: ws.row_dimensions[r].height=h

    ws.merge_cells(f"A1:{get_column_letter(TOTAL)}1")
    c=ws["A1"]; c.value="OWG 2026 Ice Dance \u2014 Rhythm Dance  |  Three-Regime Scoring Comparison"
    c.font=FONT_TITLE; c.fill=FILL_DARK; c.alignment=CTR

    ws.merge_cells(f"A2:{get_column_letter(TOTAL)}2")
    c=ws["A2"]
    fl=f"{flagged_pos} = {flagged_name} ({flagged_cc})" if flagged_pos else "No Tier 2 judge (R2 = R1)"
    r2_desc = f"OSNR-Adjusted / {flagged_pos} Removed" if flagged_pos else "No Tier 2 Flag (R2 = R1)"
    c.value=f"Official result (Regime 1) vs. Raw Average (Regime 0) vs. {r2_desc}  |  Flagged judge: {fl}"
    c.font=mk_font(10,italic=True); c.fill=FILL_GRAY; c.alignment=CTR

    def sec_hdr(c1,c2,title,fill):
        if c2>c1: ws.merge_cells(start_row=3,start_column=c1,end_row=3,end_column=c2)
        cell=ws.cell(row=3,column=c1); cell.value=title; cell.font=mk_font(10,bold=True); cell.fill=fill; cell.alignment=CTR
        for cc in range(c1,c2+1): ws.cell(row=3,column=cc).fill=fill

    sec_hdr(1,5,"ELEMENT DATA",FILL_GRAY); sec_hdr(CJ1,CJ9,"RAW JUDGE GOE SCORES",FILL_BLUE)
    sec_hdr(CR1G,CR1S,"REGIME 1: ISU Trimmed Mean (Official)",FILL_GREEN)
    sec_hdr(CR0G,CR0S,"REGIME 0: Raw Panel Average (No Trimming)",FILL_YELLOW)
    if flagged_pos:
        lname2=flagged_name.split()[-1] if flagged_name else "Removed"
        r2_hdr=f"REGIME 2: OSNR-Adjusted ({flagged_pos}/{lname2} Removed)"
    else:
        r2_hdr="REGIME 2: No Tier 2 Flag — Same as Regime 1"
    sec_hdr(CR2G,CR2S,r2_hdr,FILL_SALMON)

    def ch(col,title,fill):
        c=ws.cell(row=4,column=col); c.value=title; c.font=FONT_BOLD; c.fill=fill; c.alignment=CTR

    ch(1,"Rank\n(R1)",FILL_GRAY);ch(2,"Team",FILL_GRAY);ch(3,"NOC",FILL_GRAY);ch(4,"Element",FILL_GRAY);ch(5,"Base\nValue",FILL_GRAY)
    for i,jpos in enumerate(j_positions): ch(CJ1+i,jpos,FILL_BLUE)
    ch(CR1G,"Panel\nGOE (R1)",FILL_GREEN);ch(CR1E,"Elem\nScore (R1)",FILL_GREEN);ch(CR1T,"TES (R1)",FILL_GREEN);ch(CR1P,"PCS (R1)",FILL_GREEN);ch(CR1S,"TSS (R1)",FILL_GREEN)
    ch(CR0G,"Panel\nGOE (R0)",FILL_YELLOW);ch(CR0E,"Elem\nScore (R0)",FILL_YELLOW);ch(CR0T,"TES (R0)",FILL_YELLOW);ch(CR0P,"PCS (R0)",FILL_YELLOW);ch(CR0S,"TSS (R0)",FILL_YELLOW)
    ch(CR2G,"Panel\nGOE (R2)",FILL_SALMON);ch(CR2E,"Elem\nScore (R2)",FILL_SALMON);ch(CR2T,"TES (R2)",FILL_SALMON);ch(CR2P,"PCS (R2)",FILL_SALMON);ch(CR2S,"TSS (R2)",FILL_SALMON)

    ws.merge_cells("A5:E5")
    c=ws["A5"]; c.value="Judge Names \u2192"; c.font=FONT_JUDG; c.fill=FILL_BLUE; c.alignment=CTR
    for i,jpos in enumerate(j_positions):
        c=ws.cell(row=5,column=CJ1+i); c.value=judges[jpos]["name"]; c.font=FONT_JUDG; c.fill=FILL_BLUE; c.alignment=CTR
    for col in range(CR1G,TOTAL+1): ws.cell(row=5,column=col).fill=FILL_WHITE
    for r2 in range(1,6):
        for col in range(1,TOTAL+1): ws.cell(row=r2,column=col).border=sb(col)

    cur_row=6
    for entry in entries:
        n=len(entry["elements"]); r_start=cur_row; r_end=cur_row+n-1; rk=entry["rank"]
        alt=solid_fill("F2F2F2") if rk%2==1 else FILL_WHITE
        for ei,el in enumerate(entry["elements"]):
            r=cur_row+ei; ws.row_dimensions[r].height=15
            def sc2(col,val,fill=None,fmt=None):
                c=ws.cell(row=r,column=col); c.value=val; c.alignment=CTR
                c.fill=fill or alt; c.font=FONT_DATA
                if fmt: c.number_format=fmt
                c.border=sb(col)
            sc2(CE,el["element_code"]); sc2(CB,el["base_value"],fmt="0.00")
            for j_idx,jpos in enumerate(j_positions):
                ci=CJ1+j_idx
                if jpos in el["goe_by_pos"]:
                    gi=el["goe_by_pos"][jpos]["goe"]
                    c=ws.cell(row=r,column=ci,value=gi); c.fill=goe_fill(gi); c.font=FONT_DATA; c.alignment=CTR; c.border=sb(ci)
                else: ws.cell(row=r,column=ci).fill=FILL_GRAY
            sc2(CR1G,el["r1_goe"],FILL_GREEN,"0.00"); sc2(CR1E,el["r1_elem"],FILL_GREEN,"0.00")
            sc2(CR0G,el["r0_goe"],FILL_YELLOW,"0.00"); sc2(CR0E,el["r0_elem"],FILL_YELLOW,"0.00")
            sc2(CR2G,el["r2_goe"],FILL_SALMON,"0.00"); sc2(CR2E,el["r2_elem"],FILL_SALMON,"0.00")
            fills_m={CR1T:FILL_GREEN,CR1P:FILL_GREEN,CR1S:FILL_GREEN,CR0T:FILL_YELLOW,CR0P:FILL_YELLOW,CR0S:FILL_YELLOW,CR2T:FILL_SALMON,CR2P:FILL_SALMON,CR2S:FILL_SALMON}
            for col,fill in fills_m.items(): ws.cell(row=r,column=col).fill=fill; ws.cell(row=r,column=col).border=sb(col)
            for col in [CR,CT,CN]: ws.cell(row=r,column=col).fill=FILL_GRAY; ws.cell(row=r,column=col).border=sb(col)

        def ms(col,val,fill,fmt="0.00",bold=False):
            if n>1: ws.merge_cells(start_row=r_start,start_column=col,end_row=r_end,end_column=col)
            c=ws.cell(row=r_start,column=col); c.value=val; c.fill=fill; c.alignment=CTR
            c.font=FONT_BOLD if bold else FONT_DATA
            if fmt: c.number_format=fmt; c.border=sb(col)
        ms(CR,rk,FILL_GRAY,"0",bold=True); ms(CT,entry["team_name"],FILL_GRAY,None,bold=True); ms(CN,entry["noc"],FILL_GRAY,None)
        ms(CR1T,entry["tes_r1"],FILL_GREEN,bold=True); ms(CR1P,entry["pcs_r1"],FILL_GREEN,bold=True); ms(CR1S,entry["tss_r1"],FILL_GREEN,bold=True)
        ms(CR0T,entry["tes_r0"],FILL_YELLOW,bold=True); ms(CR0P,entry["pcs_r0"],FILL_YELLOW,bold=True); ms(CR0S,entry["tss_r0"],FILL_YELLOW,bold=True)
        ms(CR2T,entry["tes_r2"],FILL_SALMON,bold=True); ms(CR2P,entry["pcs_r2"],FILL_SALMON,bold=True); ms(CR2S,entry["tss_r2"],FILL_SALMON,bold=True)
        cur_row+=n

    sr=cur_row+2
    ws.merge_cells(f"A{sr}:{get_column_letter(TOTAL)}{sr}")
    c=ws[f"A{sr}"]; c.value="FINAL RANKINGS COMPARISON"; c.font=FONT_WHBOLD; c.fill=FILL_DARK; c.alignment=CTR; ws.row_dimensions[sr].height=22

    nr=sr+1
    ws.merge_cells(f"A{nr}:{get_column_letter(TOTAL)}{nr}")
    c=ws[f"A{nr}"]
    if flagged_pos:
        cf_w=data["lojo_summary"].get(flagged_pos,{}).get("cf_winner","")
        c.value=f"{flagged_pos} = {flagged_name} ({flagged_cc}) \u2014 OSNR flagged  |  LOJO analysis: removing {flagged_pos} reverses gold/silver ({cf_w} would rank 1st)"
    else:
        c.value="No judge meets Tier 2 criteria (winner change \u2227 p\u22640.001) \u2014 R2 = R1 (no removal performed)"
    c.font=mk_font(10,italic=True); c.fill=FILL_GRAY; c.alignment=LFT; ws.row_dimensions[nr].height=28

    hr=nr+1; ws.row_dimensions[hr].height=28
    for col,hdr in [(1,"R1\nRank"),(2,"Team"),(3,"NOC"),(4,"R1\nTSS"),(5,"R0\nRank"),(6,"R0\nTSS"),(7,"R0 Rank\nChange"),(8,"R2\nRank"),(9,"R2\nTSS"),(10,"R2 Rank\nChange")]:
        c=ws.cell(row=hr,column=col); c.value=hdr; c.font=FONT_WHBOLD; c.fill=FILL_DARK; c.alignment=CTR; c.border=std_border()

    for entry in entries:
        r=hr+entry["rank"]
        d0=entry["rank_r0"]-entry["rank"]; d2=entry["rank_r2"]-entry["rank"]
        s0=(f"+{abs(d0)}" if d0<0 else f"-{d0}") if d0!=0 else "\u2014"
        s2=(f"+{abs(d2)}" if d2<0 else f"-{d2}") if d2!=0 else "\u2014"
        rf=solid_fill("F9F9F9") if entry["rank"]%2==1 else FILL_WHITE
        def sc3(col,val,fmt=None,fill=None,fnt=None):
            c=ws.cell(row=r,column=col); c.value=val; c.alignment=CTR
            c.fill=fill or rf; c.font=fnt or FONT_DATA
            if fmt: c.number_format=fmt; c.border=std_border()
        sc3(1,entry["rank"],"0"); sc3(2,entry["team_name"]); sc3(3,entry["noc"])
        sc3(4,entry["tss_r1"],"0.00"); sc3(5,entry["rank_r0"],"0"); sc3(6,entry["tss_r0"],"0.00")
        r0f=solid_fill("FFEB9C") if d0!=0 else rf; r0c="006400" if d0<0 else ("CC0000" if d0>0 else "808080")
        sc3(7,s0,fill=r0f,fnt=Font(name="Calibri",size=10,bold=(d0!=0),color=r0c))
        sc3(8,entry["rank_r2"],"0"); sc3(9,entry["tss_r2"],"0.00")
        r2f=solid_fill("FCE4D6") if d2!=0 else rf; r2c="006400" if d2<0 else ("CC0000" if d2>0 else "808080")
        sc3(10,s2,fill=r2f,fnt=Font(name="Calibri",size=10,bold=(d2!=0),color=r2c))

    ws.freeze_panes="F6"


def build_tab_judge_statistics(ws, data):
    judges=data["judges"]; j_positions=data["judge_positions"]; judge_stats=data["judge_stats"]
    pairwise_agg=data["pairwise_agg"]; max_bias_rows=data["max_bias_rows"]; lojo_sum=data["lojo_summary"]

    cw={1:8,2:28,3:8,4:12,5:10,6:10,7:14,8:14,9:8,10:28,11:10,12:12,13:28,14:12}
    for cn,w in cw.items(): ws.column_dimensions[get_column_letter(cn)].width=w

    ws.merge_cells("A1:N1")
    c=ws["A1"]; c.value="Judge Statistics \u2014 OWG 2026 Ice Dance Rhythm Dance"
    c.font=FONT_TITLE; c.fill=FILL_DARK; c.alignment=CTR; ws.row_dimensions[1].height=22

    ws.merge_cells("A2:N2")
    c=ws["A2"]; c.value="OSNR Analysis: B(j) bias test results, LOJO counterfactual, panel behavior metrics"
    c.font=mk_font(10,italic=True); c.fill=FILL_GRAY; c.alignment=CTR; ws.row_dimensions[2].height=18

    ws.row_dimensions[3].height=8

    hdrs=["Position","Name","Country","Mean GOE Dev","Bias Z-Score","Panel Corr.",
          "# Sig Pairs\n(p\u22640.01)","# Sig Pairs\n(p\u22640.001)",
          "Max B(j)","Max B(j)\nOpponent","Kendall \u03c4","Winner\nChange?","CF Winner","OSNR Status"]
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=4,column=ci); c.value=h; c.font=FONT_WHBOLD; c.fill=FILL_DARK; c.alignment=CTR; c.border=std_border()
    ws.row_dimensions[4].height=36

    def sort_key(jpos):
        return (-pairwise_agg.get(jpos,{}).get("sig01",0), jpos)

    for idx,jpos in enumerate(sorted(j_positions,key=sort_key)):
        r=5+idx; ws.row_dimensions[r].height=18
        j=judges.get(jpos,{"id":0,"name":jpos,"country":""})
        js=judge_stats.get(jpos,{})
        pa=pairwise_agg.get(jpos,{"sig01":0,"sig001":0,"max_bias":0})
        mb=max_bias_rows.get(jpos,{})
        ls=lojo_sum.get(jpos,{})
        sig01=pa.get("sig01",0); sig001=pa.get("sig001",0); max_b=pa.get("max_bias",0)
        max_b_opp=mb.get("skater_b","") if mb else ""
        wc_flag=ls.get("winner_changes",0); pc_flag=ls.get("podium_changes",0)
        cf_winner=ls.get("cf_winner","") or ""; actual_w=ls.get("actual_winner","") or ""
        cf_show=cf_winner if cf_winner!=actual_w else ""
        osnr=("Tier 2" if sig001>0 and pc_flag>0 else "Tier 1" if sig01>0 else "No flag")
        if sig001>0: row_bg=solid_fill("F4B942")
        elif sig01>0: row_bg=solid_fill("FFFFD0")
        else: row_bg=FILL_WHITE
        osnr_fill=(solid_fill("C6EFCE") if osnr=="No flag" else solid_fill("FFEB9C") if osnr=="Tier 1" else solid_fill("F4B942"))
        def sc4(col,val,fmt=None,fill=None,fnt=None,algn=CTR):
            c=ws.cell(row=r,column=col); c.value=val; c.alignment=algn
            c.fill=fill if fill is not None else row_bg; c.font=fnt or FONT_DATA
            if fmt: c.number_format=fmt; c.border=std_border()
        sc4(1,jpos); sc4(2,j["name"],algn=LFT); sc4(3,j["country"] or "\u2014")
        sc4(4,js.get("mean_goe_dev",0),fmt="+0.000;-0.000")
        sc4(5,js.get("bias_z",0),fmt="+0.00;-0.00")
        sc4(6,js.get("correlation",0),fmt="0.000")
        sc4(7,sig01,fmt="0"); sc4(8,sig001,fmt="0")
        sc4(9,max_b,fmt="0.00"); sc4(10,max_b_opp,algn=LFT)
        sc4(11,ls.get("kendall_tau",0),fmt="0.00000")
        sc4(12,"YES" if wc_flag>0 else "No",fnt=mk_font(10,bold=wc_flag>0,color=("CC0000" if wc_flag>0 else "000000")))
        sc4(13,cf_show,algn=LFT); sc4(14,osnr,fill=osnr_fill)


def build_tab_pairwise(ws, data):
    pairwise_rows=data["pairwise_rows"]; pw_totals=data["pw_totals"]
    cw={1:8,2:22,3:32,4:6,5:32,6:6,7:7,8:9,9:14,10:11,11:11,12:12,13:10,14:10,15:10}
    for cn,w in cw.items(): ws.column_dimensions[get_column_letter(cn)].width=w

    ws.merge_cells("A1:O1")
    c=ws["A1"]; c.value="Pairwise B(j) Bias Statistics \u2014 OWG 2026 Ice Dance Rhythm Dance"
    c.font=FONT_TITLE; c.fill=FILL_DARK; c.alignment=CTR; ws.row_dimensions[1].height=22

    ws.merge_cells("A2:O2")
    c=ws["A2"]; c.value="All judge-pair combinations with p \u2264 0.10 | Sorted by p-value ascending | Use column filters to explore"
    c.font=mk_font(10,italic=True); c.fill=FILL_GRAY; c.alignment=CTR; ws.row_dimensions[2].height=18

    ws.merge_cells("A3:O3")
    total,s01,s001,sbon=pw_totals
    c=ws["A3"]
    c.value=(f"Total pairs tested: {total:,} ({total//9 if total else 0} per judge \u00d7 9 judges) | "
             f"Significant at p\u22640.01: {s01 or 0} | Significant at p\u22640.001: {s001 or 0} | Bonferroni-significant: {sbon or 0}")
    c.font=mk_font(10,bold=True); c.fill=solid_fill("FFFFD0"); c.alignment=LFT; ws.row_dimensions[3].height=18
    ws.row_dimensions[4].height=8

    hdrs5=["Judge Pos","Judge Name","Team A (Rank)","Team A\nNOC","Team B (Rank)","Team B\nNOC",
           "B(j)","p-value","Sig Level","Mean Dev A","Mean Dev B","Differential","Elements A","Elements B","Direction"]
    for ci,h in enumerate(hdrs5,1):
        c=ws.cell(row=5,column=ci); c.value=h; c.font=FONT_WHBOLD; c.fill=FILL_DARK; c.alignment=CTR; c.border=std_border()
    ws.row_dimensions[5].height=28
    ws.auto_filter.ref=f"A5:{get_column_letter(len(hdrs5))}5"

    for idx,rd in enumerate(pairwise_rows):
        r=6+idx
        jpos,jname,jcc,sa,sac,sa_rank,sb2,sbc,sb_rank,bias,pval,sig01,sig001,sigbon,mda,mdb,diff,ela,elb=rd
        if sig001: sl="\u2605\u2605\u2605 p\u22640.001"; sf=solid_fill("F4B942")
        elif sig01: sl="\u2605\u2605 p\u22640.01"; sf=solid_fill("FFEB9C")
        elif pval<=0.05: sl="\u2605 p\u22640.05"; sf=solid_fill("FFFFD0")
        else: sl="p\u22640.10"; sf=FILL_WHITE
        if bias>0: dirtxt="Favors A"; dirc="006400"
        elif bias<0: dirtxt="Favors B"; dirc="CC0000"
        else: dirtxt="Neutral"; dirc="000000"
        ta=f"{sa} (#{sa_rank})"; tb=f"{sb2} (#{sb_rank})"
        for ci,val,fmt,algn,fill_ in [
            (1,jpos,None,CTR,FILL_WHITE),(2,jname,None,LFT,FILL_WHITE),(3,ta,None,LFT,FILL_WHITE),
            (4,sac,None,CTR,FILL_WHITE),(5,tb,None,LFT,FILL_WHITE),(6,sbc,None,CTR,FILL_WHITE),
            (7,bias,"0.00",CTR,FILL_WHITE),(8,pval,"0.00000",CTR,FILL_WHITE),(9,sl,None,CTR,sf),
            (10,mda,"+0.000;-0.000",CTR,FILL_WHITE),(11,mdb,"+0.000;-0.000",CTR,FILL_WHITE),
            (12,diff,"+0.000;-0.000",CTR,FILL_WHITE),(13,ela,"0",CTR,FILL_WHITE),(14,elb,"0",CTR,FILL_WHITE),
        ]:
            c=ws.cell(row=r,column=ci); c.value=val; c.fill=fill_; c.alignment=algn; c.font=FONT_DATA
            if fmt: c.number_format=fmt; c.border=std_border()
        c=ws.cell(row=r,column=15); c.value=dirtxt; c.fill=FILL_WHITE; c.alignment=CTR
        c.font=mk_font(10,color=dirc); c.border=std_border()
        ws.row_dimensions[r].height=15

    note_row=6+len(pairwise_rows)+1
    ws.merge_cells(f"A{note_row}:O{note_row}")
    c=ws[f"A{note_row}"]
    c.value="B(j) = mean peer deviation for Team A minus mean peer deviation for Team B. Positive = judge scores Team A higher relative to panel than Team B."
    c.font=mk_font(9,italic=True); c.alignment=LFT


def build_tab_lojo(ws, data):
    judges=data["judges"]; j_positions=data["judge_positions"]; entries=data["entries"]
    lojo_sum=data["lojo_summary"]; lojo_scores=data["lojo_scores"]
    n_j=len(j_positions)
    total_cols=4+n_j*3

    cw={1:6,2:28,3:5,4:9}
    for j_idx in range(n_j):
        base=5+j_idx*3; cw[base]=8; cw[base+1]=9; cw[base+2]=5
    for cn,w in cw.items(): ws.column_dimensions[get_column_letter(cn)].width=w

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c=ws["A1"]; c.value="Leave-One-Judge-Out (LOJO) Counterfactual Rankings \u2014 OWG 2026 Ice Dance Rhythm Dance"
    c.font=FONT_TITLE; c.fill=FILL_DARK; c.alignment=CTR; ws.row_dimensions[1].height=22

    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    c=ws["A2"]; c.value="Each cell shows the team's counterfactual rank if that judge were removed. Green = rank improves, Red = rank drops, White = unchanged."
    c.font=mk_font(10,italic=True); c.fill=FILL_GRAY; c.alignment=CTR; ws.row_dimensions[2].height=18

    ws.row_dimensions[3].height=8

    ws.merge_cells("A4:D4")
    c=ws["A4"]; c.value="Official Results"; c.font=FONT_BOLD; c.fill=FILL_GRAY; c.alignment=CTR

    for j_idx,jpos in enumerate(j_positions):
        base=5+j_idx*3
        ls=lojo_sum.get(jpos,{})
        tau=ls.get("kendall_tau",0) or 0; pc=ls.get("podium_changes",0) or 0
        txt=f"{jpos}: \u03c4={tau:.3f} | {'Podium change' if pc>0 else 'No podium change'}"
        ws.merge_cells(start_row=4,start_column=base,end_row=4,end_column=base+2)
        c=ws.cell(row=4,column=base); c.value=txt
        c.font=mk_font(9,bold=(pc>0),color=("CC0000" if pc>0 else "000000"))
        c.fill=solid_fill("FFFFD0") if pc>0 else FILL_LGRAY; c.alignment=CTR
    ws.row_dimensions[4].height=22

    for ci,h in [(1,"Off.\nRank"),(2,"Team"),(3,"NOC"),(4,"Official\nTSS")]:
        c=ws.cell(row=5,column=ci); c.value=h; c.font=FONT_BOLD; c.fill=FILL_GRAY; c.alignment=CTR; c.border=std_border()
    for j_idx,jpos in enumerate(j_positions):
        base=5+j_idx*3
        for offset,hdr in [(0,f"{jpos}\nCF Rank"),(1,f"{jpos}\nCF TSS"),(2,f"{jpos}\n\u0394")]:
            c=ws.cell(row=5,column=base+offset); c.value=hdr; c.font=FONT_BOLD; c.fill=FILL_BLUE; c.alignment=CTR; c.border=std_border()
    ws.row_dimensions[5].height=30
    ws.freeze_panes="E6"

    for e_idx,entry in enumerate(entries):
        r=6+e_idx; eid=entry["entry_id"]; rk=entry["rank"]
        alt=FILL_LGRAY if rk%2==1 else FILL_WHITE
        for ci,val,fmt in [(1,rk,"0"),(2,entry["team_name"],None),(3,entry["noc"],None),(4,entry["tss_r1"],"0.00")]:
            c=ws.cell(row=r,column=ci); c.value=val; c.fill=alt
            c.font=FONT_BOLD if ci==1 else FONT_DATA; c.alignment=LFT if ci==2 else CTR
            if fmt: c.number_format=fmt; c.border=std_border()
        for j_idx,jpos in enumerate(j_positions):
            base=5+j_idx*3
            ls_e=lojo_scores.get(jpos,{}).get(eid,{})
            cf_rank=ls_e.get("cf_rank"); cf_tss=ls_e.get("cf_tss"); rchg=ls_e.get("rank_change",0) or 0
            if rchg<0: cf=solid_fill("C6EFCE")
            elif rchg>0: cf=solid_fill("FCE4D6")
            else: cf=FILL_WHITE
            dtxt=(f"+{abs(rchg)}" if rchg<0 else f"-{rchg}" if rchg>0 else "0")
            for offset,val,fmt in [(0,cf_rank,"0"),(1,cf_tss,"0.00"),(2,dtxt,None)]:
                c=ws.cell(row=r,column=base+offset); c.value=val; c.fill=cf; c.alignment=CTR; c.font=FONT_DATA
                if fmt and val is not None: c.number_format=fmt; c.border=std_border()
        ws.row_dimensions[r].height=15


def build_tab_legend(ws, data):
    cw={1:8,2:20,3:25,4:16,5:50}
    for cn,w in cw.items(): ws.column_dimensions[get_column_letter(cn)].width=w

    row=[1]
    def nr(): row[0]+=1; return row[0]-1

    ws.merge_cells("A1:E1")
    c=ws["A1"]; c.value="Legend & Methodology Notes \u2014 OSNR Analysis Report"
    c.font=FONT_TITLE; c.fill=FILL_DARK; c.alignment=CTR; ws.row_dimensions[1].height=22

    def sec_hdr(title,fill=FILL_BLUE):
        row[0]+=1; r=row[0]
        ws.merge_cells(f"A{r}:E{r}")
        c=ws[f"A{r}"]; c.value=title; c.font=mk_font(10,bold=True); c.fill=fill; c.alignment=CTR; ws.row_dimensions[r].height=18
        row[0]+=1

    def lgd_row(v1,v2,v3,fill=FILL_WHITE):
        r=row[0]
        c=ws.cell(row=r,column=1); c.value=v1; c.fill=fill; c.alignment=CTR; c.font=FONT_DATA
        ws.merge_cells(f"B{r}:C{r}")
        c=ws.cell(row=r,column=2); c.value=v2; c.fill=fill; c.alignment=CTR; c.font=FONT_DATA
        ws.merge_cells(f"D{r}:E{r}")
        c=ws.cell(row=r,column=4); c.value=v3; c.fill=fill; c.alignment=LFT; c.font=FONT_DATA
        ws.row_dimensions[r].height=15; row[0]+=1

    def note_r(label,text,h=55):
        r=row[0]
        c=ws.cell(row=r,column=1); c.value=label; c.font=mk_font(10,bold=True); c.fill=FILL_GRAY; c.alignment=LTOP
        ws.merge_cells(f"B{r}:E{r}")
        c=ws.cell(row=r,column=2); c.value=text; c.font=FONT_DATA; c.fill=FILL_WHITE; c.alignment=LTOP
        ws.row_dimensions[r].height=h; row[0]+=1

    sec_hdr("Section 1: GOE Color Scale (Individual Judge Scores)")
    for gv,cn,desc in [(-5,"Deep Red","Very poor execution"),(-4,"Red","Poor"),(-3,"Light Red","Below average"),
        (-2,"Pale Red","Slightly below"),(-1,"Very Pale Red","Marginally below"),(0,"White","Average / panel mean"),
        (1,"Very Pale Green","Marginally above"),(2,"Pale Green","Slightly above"),
        (3,"Light Green","Above average"),(4,"Green","Very good"),(5,"Deep Green","Outstanding")]:
        lgd_row(gv,cn,desc,fill=PatternFill(fill_type="solid",fgColor=GOE_COLORS[gv]))

    sec_hdr("Section 2: Regime Color Coding")
    for fill_,name,desc in [
        (FILL_GRAY,"Element Data","Columns A-E: Rank, team, NOC, element code, base value"),
        (FILL_BLUE,"Raw Judge Scores","Columns J1-J9: Individual judge GOE integer scores"),
        (FILL_GREEN,"Regime 1 (R1)","ISU Official: trimmed mean (drop 1H+1L of 9 judges, avg 7)"),
        (FILL_YELLOW,"Regime 0 (R0)","Raw Mean: simple average of all 9 judges, no trimming"),
        (FILL_SALMON,"Regime 2 (R2)","OSNR-Adjusted: remove flagged judge, trim remaining 8 (drop 1H+1L, avg 6)"),
    ]:
        r=row[0]
        c=ws.cell(row=r,column=1); c.fill=fill_; c.alignment=CTR; c.font=mk_font(10,bold=True)
        ws.merge_cells(f"B{r}:C{r}")
        c=ws.cell(row=r,column=2); c.value=name; c.fill=fill_; c.font=mk_font(10,bold=True); c.alignment=CTR
        ws.merge_cells(f"D{r}:E{r}")
        c=ws.cell(row=r,column=4); c.value=desc; c.fill=fill_; c.font=FONT_DATA; c.alignment=LFT
        ws.row_dimensions[r].height=15; row[0]+=1

    sec_hdr("Section 3: LOJO Counterfactual Heatmap Color Coding")
    for fill_,name,desc in [
        (solid_fill("C6EFCE"),"Rank improved","Counterfactual rank is lower (better) than official"),
        (solid_fill("FCE4D6"),"Rank dropped","Counterfactual rank is higher (worse) than official"),
        (FILL_WHITE,"Rank unchanged","Same rank with or without this judge"),
    ]:
        r=row[0]
        c=ws.cell(row=r,column=1); c.fill=fill_; c.alignment=CTR
        ws.merge_cells(f"B{r}:C{r}")
        c=ws.cell(row=r,column=2); c.value=name; c.fill=fill_; c.font=FONT_DATA; c.alignment=CTR
        ws.merge_cells(f"D{r}:E{r}")
        c=ws.cell(row=r,column=4); c.value=desc; c.fill=fill_; c.font=FONT_DATA; c.alignment=LFT
        ws.row_dimensions[r].height=15; row[0]+=1

    sec_hdr("Section 4: Significance Level Coding (Pairwise B(j) Tab)")
    for fill_,label,desc in [
        (solid_fill("F4B942"),"\u2605\u2605\u2605 p\u22640.001","Very strong evidence - Bonferroni-class significance"),
        (solid_fill("FFEB9C"),"\u2605\u2605 p\u22640.01","Strong evidence - meets OSNR Tier 1 threshold"),
        (solid_fill("FFFFD0"),"\u2605 p\u22640.05","Moderate evidence - conventional significance"),
        (FILL_WHITE,"p\u22640.10","Weak evidence - included for completeness"),
    ]:
        r=row[0]
        c=ws.cell(row=r,column=1); c.fill=fill_; c.alignment=CTR
        ws.merge_cells(f"B{r}:C{r}")
        c=ws.cell(row=r,column=2); c.value=label; c.fill=fill_; c.font=FONT_DATA; c.alignment=CTR
        ws.merge_cells(f"D{r}:E{r}")
        c=ws.cell(row=r,column=4); c.value=desc; c.fill=fill_; c.font=FONT_DATA; c.alignment=LFT
        ws.row_dimensions[r].height=15; row[0]+=1

    sec_hdr("Section 5: Methodology Notes",FILL_DARK)
    ws.cell(row=row[0]-1,column=1).font=FONT_WHBOLD
    note_r("B(j) Statistic:",
        "B(j) measures a judge's differential scoring bias between two teams. "
        "B(j) = mean GOE deviation of Team A minus mean GOE deviation of Team B, where deviation = "
        "this judge's score minus the panel trimmed mean for each element. "
        "Positive B(j) = judge systematically scores Team A higher relative to panel consensus.")
    note_r("Permutation Test:",
        "Significance of B(j) assessed via permutation test. Judge's element-level GOE scores are "
        "randomly reassigned across the two teams, B(j) recomputed for each permutation. "
        "p-value = fraction of permuted |B(j)| >= observed |B(j)|.")
    note_r("LOJO Procedure:",
        "Leave-One-Judge-Out: for each judge J, remove J's scores and recompute TES (trimmed mean of "
        "remaining 8: drop 1H+1L, avg 6), PCS similarly, then re-rank all skaters by counterfactual TSS. "
        "Kendall tau distance measures ranking disruption caused by each judge.")
    note_r("Three Regimes:",
        "R1 (Official): ISU trimmed mean of 9 judges (drop 1H+1L, avg 7). "
        "R0 (Raw Mean): simple average of all 9 judges, no trimming. "
        "R2 (OSNR-Adjusted): remove flagged judge, trim remaining 8 (drop 1H+1L, avg 6). "
        "PCS computed analogously for each regime.")
    note_r("OSNR Decision Rule:",
        "Tier 1 flag: any judge-pair has p<=0.01 in permutation B(j) test. "
        "Tier 2 flag (OSNR-positive): Tier 1 met AND at least one judge's removal changes any "
        "top-3 podium position (podium_changes>0 in LOJO analysis). "
        "Tier 2 triggers full OSNR-adjusted regime (R2) for official results review.")
    note_r("GOE Increment Factor:",
        "For each element: increment = official_panel_goe_points / official_trimmed_mean_integer. "
        "This converts trimmed/raw mean integer GOE to score points (element-specific per ISU SoV for Ice Dance).")


def generate_filename(event_info, base_dir):
    discipline = event_info.get("discipline","Unknown").replace(" ","")[:10]
    segment    = event_info.get("segment","Unknown").replace(" ","")[:10]
    dt = event_info.get("datetime_local","")
    year_str   = dt[:4] if dt and len(dt)>=4 else ""
    comp_name  = event_info.get("competition_name","")
    if "OWG" in comp_name or "Olympic" in comp_name: prefix=f"OWG{year_str}"
    elif "WC" in comp_name or "World" in comp_name: prefix=f"WC{year_str}"
    elif "GP" in comp_name or "Grand Prix" in comp_name: prefix=f"GP{year_str}"
    else: prefix=f"Event{event_info['event_id']}"
    disc_map={"IceDance":"ID","MenSingles":"MS","LadiesSingles":"LS","Pairs":"PR"}
    seg_map={"RhythmDance":"RD","FreeDance":"FD","ShortProgram":"SP","FreeSkating":"FS"}
    sd=disc_map.get(discipline,discipline[:3] if len(discipline)>=3 else discipline)
    ss=seg_map.get(segment,segment[:3] if len(segment)>=3 else segment)
    return os.path.join(base_dir, f"{prefix}_{sd}_{ss}_OSNR.xlsx")


def main():
    args=parse_args()
    event_id=args.event_id; db_path=args.db
    if not os.path.exists(db_path):
        print(f"ERROR: Database not found: {db_path}",file=sys.stderr); sys.exit(1)
    print(f"Loading event {event_id} from {db_path} ...")
    conn=sqlite3.connect(db_path)
    data=load_event_data(conn,event_id)
    conn.close()
    ev=data["event_info"]
    print(f"  Event: {ev['discipline']} {ev['segment']} ({ev['competition_name']})")
    print(f"  Entries: {len(data['entries'])}, Judges: {len(data['judge_positions'])}")
    print(f"  Flagged judge: {ev.get('flagged_judge_pos','none')}")
    if args.output: out_path=args.output
    else: out_path=generate_filename(ev,os.path.dirname(db_path))
    print(f"\nBuilding workbook: {out_path}")
    wb=Workbook()
    ws_summary=wb.active; ws_summary.title="Summary"
    ws_raw=wb.create_sheet("Raw Scores")
    ws_regime=wb.create_sheet("Regime Comparison")
    ws_jstats=wb.create_sheet("Judge Statistics")
    ws_pairwise=wb.create_sheet("Pairwise B(j)")
    ws_lojo=wb.create_sheet("LOJO Counterfactual")
    ws_legend=wb.create_sheet("Legend")
    print("  Building Tab 1: Summary ..."); build_tab_summary(ws_summary,data)
    print("  Building Tab 2: Raw Scores ..."); build_tab_raw_scores(ws_raw,data)
    print("  Building Tab 3: Regime Comparison ..."); build_tab_regime_comparison(ws_regime,data)
    print("  Building Tab 4: Judge Statistics ..."); build_tab_judge_statistics(ws_jstats,data)
    print("  Building Tab 5: Pairwise B(j) ..."); build_tab_pairwise(ws_pairwise,data)
    print("  Building Tab 6: LOJO Counterfactual ..."); build_tab_lojo(ws_lojo,data)
    print("  Building Tab 7: Legend ..."); build_tab_legend(ws_legend,data)
    print(f"\nSaving {out_path} ...")
    wb.save(out_path)
    file_size=os.path.getsize(out_path)
    print(f"\n{'='*60}")
    print(f"VERIFICATION SUMMARY")
    print(f"{'='*60}")
    print(f"Output file: {out_path}")
    print(f"File size: {file_size:,} bytes ({file_size/1024:.1f} KB)")
    print()
    print("Tab names and row counts:")
    from openpyxl import load_workbook
    wb_verify=load_workbook(out_path,read_only=True,data_only=True)
    for sheet in wb_verify.sheetnames:
        ws_v=wb_verify[sheet]
        print(f"  {sheet}: {ws_v.max_row} rows x {ws_v.max_column} cols")
    wb_verify.close()
    print()
    print("Key spot-checks:")
    conn2=sqlite3.connect(db_path)
    c2=conn2.cursor()
    c2.execute("SELECT SUM(is_significant_01) FROM pairwise_judge_statistics WHERE event_id=? AND judge_position='J5'",(event_id,))
    j5_sig01=c2.fetchone()[0] or 0
    print(f"  Tab 4 J5 sig pairs (p<=0.01): {j5_sig01}")
    c2.execute("SELECT COUNT(*) FROM pairwise_judge_statistics WHERE event_id=? AND p_value<=0.10",(event_id,))
    pw_rows=c2.fetchone()[0]
    print(f"  Tab 5 total rows (p<=0.10): {pw_rows}")
    c2.execute("SELECT winner_changes,cf_winner_name FROM lojo_event_summary WHERE event_id=? AND judge_position='J5'",(event_id,))
    j5_lojo=c2.fetchone()
    if j5_lojo:
        wc,cfw=j5_lojo
        print(f"  Tab 6 J5 winner change: {'YES' if wc else 'No'} (CF winner: {cfw})")
    conn2.close()
    print(f"{'='*60}")
    print("Done.")

if __name__=="__main__":
    main()
